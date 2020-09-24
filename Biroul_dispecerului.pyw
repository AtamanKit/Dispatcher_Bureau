from PySide2.QtGui import QIcon, QBrush, QColor, QFont, QPainter, QPixmap, QRegExpValidator, QMovie
from PySide2.QtCore import QAbstractTableModel, Qt, QRegExp, QSize
from PySide2.QtWidgets import QMdiSubWindow, QLineEdit, QMainWindow, QApplication, QMdiArea, \
    QDesktopWidget, QMenu, QAction, QComboBox, QLabel, QFrame, QVBoxLayout, QHBoxLayout, QSplitter, \
    QTableView, QAbstractItemView, QDialog, QCompleter, QGridLayout, QPushButton, QFileDialog, \
    QMessageBox, QProgressBar, QSplashScreen, QWidget, QTextEdit, QCheckBox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PySide2.QtCharts import QtCharts
from qtwidgets import PasswordEdit
from pymongo import MongoClient
import sys
import pandas as pd
import numpy as np
import re
import datetime
import random
import webbrowser
import os.path
import shutil
import json

class tabSubWindow(QMdiSubWindow):
    def __init__(self):
        QMdiSubWindow.__init__(self)
    # def closeEvent(self, event):
    #     mainWindow.nameRegClose(self)
    #     # print("test")

class tabDecSubWd(QMdiSubWindow):
    def __init__(self):
        QMdiSubWindow.__init__(self)
    def closeEvent(self, event):
        pass
        # mainWindow.nameRegClose(self)
        # print("test")

#Formez Model/View
class TableModel(QAbstractTableModel):

    def __init__(self, data, headerdata):
        super(TableModel, self).__init__()
        self._data = data
        self.headerdata = headerdata

    def data(self, index, role):
        if role == Qt.DisplayRole:
            value = self._data.iloc[index.row(), index.column()]
            # if isinstance(value, int):
            #     return str(value)
            # else:
            return value

        if role == Qt.TextAlignmentRole:
            value = self._data.iloc[index.row(), index.column()]
            # matched = re.match("\d\d.\d\d.\d\d\d\d \d\d:\d\d", str(value))
            # matchedShort = re.match("\d\d.\d\d.\d\d \d\d:\d\d", str(value))
            # matchedDelta = re.match("\d:\d\d:\d\d", str(value))
            # matchedLa = re.search("La ", str(value))
            # matchedInr = re.search("\d\d.\d\d.\d\d", str(value))
            # matchedAdm = re.search("admis", str(value))
            # matchedFider = re.search("Fider", str(value))
            # matchedFormatia = re.search("Formatia:", str(value))

            # if len(str(value)) <= 9 or bool(matched) or bool(matchedShort) or bool(matchedLa) or \
            #         str(value) == "Nu s-a lucrat" or str(value) == "Nou inregistrata" \
            #         or str(value) == "Programat" or str(value) == "Fara deconectari" \
            #         or str(value) == "Neprogramat" or bool(matchedDelta) \
            #         or str(value) == "/ ... /" or str(value) == "Intrerupere"\
            #         or bool(matchedInr) or bool(matchedAdm) or str(value) == "-//-" \
            #         or bool(matchedFider) or bool(matchedFormatia):
            #     return Qt.AlignCenter
            if str(value):
                return Qt.AlignCenter
        if role == Qt.BackgroundRole:
            value = self._data.iloc[index.row(), index.column()]
            # value_minus = self._data.iloc[index.row(), index.column() - 1]
            matched = re.match("\d\d.\d\d.\d\d\d\d \d\d:\d\d", str(value))
            matchedShort = re.match("\d\d.\d\d.\d\d \d\d:\d\d", str(value))
            matchedDelta = re.match("\d:\d\d:\d\d", str(value))
            matchedPT = re.search("PT\d", str(value))
            matchedLa = re.search("La ", str(value))
            matchedAdm = re.search("admis", str(value))
            matchedGr = re.search("gr. ", str(value))
            matchedFormatia = re.search("Formatia:", str(value))
            matchedExcl = re.search("!", str(value))
            matchedExec = re.search("Executat", str(value))
            if str(value) == "Nou inregistrata" or str(value) == "Neexecutat":
                # print(str(valuePlus))
                return QColor(179, 48, 48)
            elif bool(matched) or str(value) == "/ ... /" or bool(matchedAdm):
                return QColor(194, 199, 40)
            elif bool(matchedShort):
                return QColor(80, 80, 80)
            elif str(value) == "Nu s-a lucrat":
                return QColor(13, 158, 163)
            elif str(value) == "/ --- /":
                return QColor(13, 158, 163)
            elif str(value) == "Programat":
                return QColor(150, 200, 175)
            elif str(value) == "Neprogramat":
                return QColor(209, 169, 169)
            elif str(value) == "UN":
                return QColor(180, 60, 60)
            elif str(value) == "FL":
                return QColor(87, 107, 148)
            elif str(value) == "GL":
                return QColor(150, 150, 150)
            elif str(value) == "RS":
                return QColor(130, 100, 200)
            elif bool(matchedDelta):
                return QColor(70, 122, 78)
            elif bool(matchedLa):
                return QColor(79, 74, 128)
            elif bool(matchedPT):
                return QColor(51, 105, 143)
            elif bool(matchedGr) and not bool(matchedFormatia):
                return QColor(158, 81, 81)
            elif str(value) == "Intrerupere":
                return QColor(46, 130, 68)
            elif bool(matchedExcl):
                return QColor(13, 110, 106)
            elif bool(matchedExec):
                return QColor(68, 112, 71)
        if role == Qt.ForegroundRole:
            value = self._data.iloc[index.row(), index.column()]
            matchedShort = re.match("\d\d.\d\d.\d\d \d\d:\d\d", str(value))
            # matchedDelta = re.match("\d:\d\d:\d\d", str(value))
            if str(value) == "Nou inregistrata" or bool(matchedShort) or str(value) == "Neexecutat":
                return QColor(255, 255, 255)
            # elif str(value) == "Programat":
            #     return QColor(255, 255, 255)
        if role == Qt.FontRole:
            value = self._data.iloc[index.row(), index.column()]
            # matched = re.match("\d\d.\d\d.\d\d\d\d \d\d:\d\d", str(value))
            matchedShort = re.match("\d\d.\d\d.\d\d \d\d:\d\d", str(value))
            matchedDelta = re.match("\d:\d\d:\d\d", str(value))
            if str(value) == "Nou inregistrata" or str(value) == "Nu s-a lucrat" \
                    or bool(matchedShort) or bool(matchedDelta):
                return QFont("Calibri", 14, QFont.Bold)

            else:
                return QFont("Calibri", 14, QFont.Bold)


    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

    def headerData(self, section, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.headerdata[section]
        if role == Qt.BackgroundRole:
            return QColor(66, 135, 245)
        if role == Qt.ForegroundRole:
            return QColor(0, 0, 0)
        if role == Qt.FontRole:
            return QFont("Calibri", 10, QFont.Bold)
        if role == Qt.DecorationRole:
            return QColor(66, 135, 245)

        # if orientation == Qt.Vertical and role == Qt.DisplayRole:
        #     return self.headerdataH[section]
        # if role == Qt.BackgroundRole:
        #     return QColor(66, 135, 245)
        # if role == Qt.ForegroundRole:
        #     return QColor(0, 0, 0)
        # if role == Qt.FontRole:
        #     return QFont("Calibri", 10, QFont.Bold)
        # if role == Qt.DecorationRole:
        #     return QColor(66, 135, 245)

class TableModelII(QAbstractTableModel):

    def __init__(self, data, headerdata):
        super(TableModelII, self).__init__()
        self._data = data
        self.headerdata = headerdata

    def data(self, index, role):
        if role == Qt.DisplayRole or role == Qt.EditRole:
            value = self._data.iloc[index.row(), index.column()]
            return value

        if role == Qt.TextAlignmentRole:
            # value = self._data.iloc[index.row(), index.column()]
            return Qt.AlignCenter
        if role == Qt.BackgroundRole:
            value = self._data.iloc[index.row(), index.column()]
            if str(value) == "UN":
                return QColor(180, 60, 60)
            elif str(value) == "FL":
                return QColor(87, 107, 148)
            elif str(value) == "GL":
                return QColor(150, 150, 150)
            elif str(value) == "RS":
                return QColor(130, 100, 200)
        if role == Qt.FontRole:
            return QFont("Calibri", 14, QFont.Bold)

    def flags(self, index):
        flag = super(TableModelII, self).flags(index)
        return flag | Qt.ItemIsEditable

    def setData(self, index, value, role):
        if role == Qt.EditRole:
            self._data.iloc[index.row(), index.column()] = value
            # self.dataChanged.emit(index, index)
            mainWindow.dtContrDecZl(self)
            self.wsNrSol.cell(row=index.row()+5, column=index.column()+1).value = value
            self.wsNrSol.cell(row=index.row()+5, column=index.column()+1).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wbDecZil.save(self.rapFile)
            return True

    def destLoad(self):
        self.myDest = os.path.abspath(".") + "/Bundle/Destination.xlsx"
        self.wbDest = load_workbook(self.myDest)
        self.wsDest = self.wbDest["Destination"]


    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

    def headerData(self, section, orientation, role):
        if orientation == Qt.Horizontal and role == Qt.DisplayRole:
            return self.headerdata[section]
        if role == Qt.BackgroundRole:
            return QColor(66, 135, 245)
        if role == Qt.ForegroundRole:
            return QColor(0, 0, 0)
        if role == Qt.FontRole:
            return QFont("Calibri", 10, QFont.Bold)
        if role == Qt.DecorationRole:
            return QColor(66, 135, 245)

class TableModelIII(QAbstractTableModel):

    def __init__(self, data):
        super(TableModelIII, self).__init__()
        self._data = data
        # self.headerdata = headerdata

    def data(self, index, role):
        if role == Qt.DisplayRole or role == Qt.EditRole:
            value = self._data.iloc[index.row(), index.column()]
            return value

        if role == Qt.BackgroundRole:
            value = self._data.iloc[index.row(), index.column()]
            if str(value) == "UN":
                return QColor(180, 60, 60)
            elif str(value) == "FL":
                return QColor(87, 107, 148)
            elif str(value) == "GL":
                return QColor(150, 150, 150)
            elif str(value) == "RS":
                return QColor(130, 100, 200)
        if role == Qt.FontRole:
            return QFont("Calibri", 14, QFont.Bold)

        if role == Qt.TextAlignmentRole:
            value = self._data.iloc[index.row(), index.column()]
            matched = re.match("\d", str(value))
            if bool(matched) or \
                    str(value) == "UN" or str(value) == "FL" \
                    or str(value) == "GL" or str(value) == "RS":
                return Qt.AlignCenter

    def flags(self, index):
        flag = super(TableModelIII, self).flags(index)
        return flag | Qt.ItemIsEditable

    def setData(self, index, value, role):
        if role == Qt.EditRole:
            self._data.iloc[index.row(), index.column()] = value
            self.dataChanged.emit(index, index)
            mainWindow.dtContrDecZl(self)
            self.wsNrSol.cell(row=index.row()+13, column=index.column()+1).value = value
            self.wbDecZil.save(self.rapFile)
            return True

    def destLoad(self):
        self.myDest = os.path.abspath(".") + "/Bundle/Destination.xlsx"
        self.wbDest = load_workbook(self.myDest)
        self.wsDest = self.wbDest["Destination"]

    def rowCount(self, index):
        return self._data.shape[0]

    def columnCount(self, index):
        return self._data.shape[1]

#Folosesc acest clas pentru a ma putea folosi de focus out, controlez oara QLineEdit
class DateLineEdit(QLineEdit):

    def __init__(self):
        QLineEdit.__init__(self)

    # def focusInEvent(self, e):
    #     QLineEdit.focusInEvent(self, e)
    #     self.initText = self.text()
    #     self.setText("")

    def focusOutEvent(self, e):
        QLineEdit.focusOutEvent(self, e)
        matchedShort = re.match("\d\d.\d\d.\d\d \d\d:\d\d", self.text())
        if not bool(matchedShort):
            mainWindow.msSecCall(self,'Formatul necesar pentru data si ora este: \n zz.ll.aa hh:mm')
            self.setFocus()

#Folosesc acest clas pentru a ma putea folosi de focus in QLineEdit
class MyLineEdit(QLineEdit):

    def __init__(self):
        QLineEdit.__init__(self)

    def focusInEvent(self, e):
        QLineEdit.focusInEvent(self, e)
        self.initText = self.text()
        self.setText("")

    def focusOutEvent(self, e):
        QLineEdit.focusOutEvent(self, e)
        if self.text() == "":
            self.setText(self.initText)

# #Clasul loading
# class loadWidow(QMainWindow):


#Clasul principal
class mainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        # Parent Window create

        self.myMidi = QMdiArea()
        self.setCentralWidget(self.myMidi)
        self.setWindowTitle('Biroul dispecerului')
        self.setWindowIcon(QIcon('ataman_logo'))
        self.setGeometry(100, 100, 1250, 720)
        # self.myWindow.setStyleSheet('background-color: #6e6e6e')
        self.myMidi.setBackground(QBrush(QColor(100, 100, 100)))

        # qtRectangle = self.frameGeometry()
        # centerPoint = QDesktopWidget().availableGeometry().center()
        # qtRectangle.moveCenter(centerPoint)
        # self.move(qtRectangle.topLeft())
        self.showMaximized()

        # Menu Create

        bar = self.menuBar()
        fileMenu = bar.addMenu('File')

        newMen = QMenu('New', self)

        autSub = QAction('Autorizatie', self)
        newMen.addAction(autSub)
        fileMenu.addMenu(newMen)
        autSub.triggered.connect(self.alTrig)

        disSub = QAction('Dispozitie', self)
        newMen.addAction(disSub)
        fileMenu.addMenu(newMen)
        disSub.triggered.connect(self.dsTrig)

        decSub = QAction('Deconectare neplanificata', self)
        newMen.addAction(decSub)
        fileMenu.addMenu(newMen)
        decSub.triggered.connect(self.decTrig)

        deranjSub = QAction('Deranjament', self)
        newMen.addAction(deranjSub)
        fileMenu.addMenu(newMen)
        deranjSub.triggered.connect(self.deranjTrig)

        setMenu = QAction("Setari", self)
        fileMenu.addAction(setMenu)
        setMenu.triggered.connect(self.setTrig)

        fileMenu.addSeparator()

        exitMenu = QAction("Exit", self)
        fileMenu.addAction(exitMenu)
        # fileMenu.addMenu(exitMenu)
        exitMenu.triggered.connect(self.close)

        veziMen = bar.addMenu('Vezi')
        veziAl = QAction('Autorizatii si dispozitii', self)
        veziMen.addAction(veziAl)
        veziAl.triggered.connect(self.centrAlPop)
        veziDec = QAction('Raport deconectari', self)
        veziMen.addAction(veziDec)
        veziDec.triggered.connect(self.decTabWindow)
        veziDeranj = QAction('Deranjamente', self)
        veziMen.addAction(veziDeranj)
        veziDeranj.triggered.connect(self.deranjPop)

        anMen = bar.addMenu('Analiza')
        anDec = QAction("Lunara", self)
        anMen.addAction(anDec)
        anDec.triggered.connect(self.decAnalizaLun)
        anualDec = QAction("Anuala", self)
        anMen.addAction(anualDec)
        anualDec.triggered.connect(self.decAnalizaAn)

        autMen = bar.addMenu('LogIn/Out')
        autIn = QAction("Log In", self)
        autMen.addAction(autIn)
        autMen.triggered.connect(self.intTrig)
        autOut = QAction("Log Out", self)
        autMen.addAction(autOut)
        autOut.triggered.connect(self.logOut)

        self.deranjControlPop = False
        self.tabWindowControl = False
        self.tabDecControl = False
        self.anControl = False
        self.passControl = False
        #Controlez la terminarea lucrarilor ca nu e deschis registru autorizatii
        self.erContrAl = False

        self.destLoad()
        if self.wsDest.cell(row=1, column=2).value == None:
            self.msCall("destinatia pentru Registru AUTORIZATII excel!")
            self.setTrig()
        elif self.wsDest.cell(row=2, column=2).value == None:
            self.msCall("destinatia pentru SAIDI excel!")
            self.setTrig()
        elif self.wsDest.cell(row=3, column=2).value == None:
            self.msCall("destinatia pentru RAPORTUL PDJT excel!")
            self.setTrig()

        self.intTrig()

    #Incarc mongoDB
    def loadMongo(self):
        try:
            self.client = MongoClient("mongodb+srv://PdjtUn:123pdj34@red-nord.lhwnm.mongodb.net/test?"
                     "retryWrites=true&w=majority")
        except:
            self.msSecCall("Lipseste legatura cu internetul!")
    # def loadMongoGen(self):
    #     self.loadMongo()
    #     self.db = self.client.General

    def loadOficii(self):
        self.ofList = []
        self.ofListAbr = []
        with open("Bundle/Oficii.json", "r") as f:
            oficiiDict = json.load(f)

        for i in oficiiDict:
            self.ofList.append(i["name"])
            self.ofListAbr.append(i["abr"])

    def loadMongoUN(self):
        self.loadMongo()
        self.db = self.client.Ungheni

    def loadAng(self):
        if self.ofCombo.currentText() == "Ungheni":
            self.angajati = self.db.angajati_un
        if self.ofCombo.currentText() == "Falesti":
            self.angajati = self.db.angajati_fl
        if self.ofCombo.currentText() == "Glodeni":
            self.angajati = self.db.angajati_gl
        if self.ofCombo.currentText() == "Riscani":
            self.angajati = self.db.angajati_rs

        self.uList = []
        for i in self.angajati.find():
            if i["gr_ts"] == "5":
                self.uList.append(i["name"])
        self.uList.sort()
        self.uCombo.clear()
        self.uCombo.addItems(self.uList)
        # self.uCombo.setCurrentText("Alege:")

    #Abrevieri oficii
    def abrOficii(self):
        if self.ofCombo.currentText() == "Ungheni":
            self.ofVar = "UN"
        if self.ofCombo.currentText() == "Falesti":
            self.ofVar = "FL"
        if self.ofCombo.currentText() == "Glodeni":
            self.ofVar = "GL"
        if self.ofCombo.currentText() == "Riscani":
            self.ofVar = "RS"
    #Incarc informatia PT excel autorizati, dispozitie, deconectare neplanificata
    def loadPt(self):
        if self.ofCombo.currentText() == "Ungheni":
            self.wbPt = load_workbook("Bundle/Ungheni/PT_Ungheni.xlsx")
            if not self.decControl:
                self.angajati = self.db.angajati_un
            if self.deranjControl:
                self.sector = open("Bundle/Ungheni/Sector.txt")
                self.f10kv = open("Bundle/Ungheni/Lista_f_10kV.txt")
        if self.ofCombo.currentText() == "Falesti":
            self.wbPt = load_workbook("Bundle/Falesti/PT_Falesti.xlsx")
            if not self.decControl:
                self.angajati = self.db.angajati_fl
            if self.deranjControl:
                self.sector = open("Bundle/Falesti/Sector.txt")
        if self.ofCombo.currentText() == "Glodeni":
            self.wbPt = load_workbook("Bundle/Glodeni/PT_Glodeni.xlsx")
            if not self.decControl:
                self.angajati = self.db.angajati_gl
                # self.wbAng = load_workbook("Bundle/Glodeni/Angajati.xlsx")
            if self.deranjControl:
                # self.wbAng = load_workbook("Bundle/Glodeni/Angajati.xlsx")
                self.sector = open("Bundle/Glodeni/Sector.txt")
        if self.ofCombo.currentText() == "Riscani":
            self.wbPt = load_workbook("Bundle/Riscani/PT_Riscani.xlsx")
            if not self.decControl:
                self.angajati = self.db.angajati_rs
                # self.wbAng = load_workbook("Bundle/Riscani/Angajati.xlsx")
            if self.deranjControl:
                # self.wbAng = load_workbook("Bundle/Riscani/Angajati.xlsx")
                self.sector = open("Bundle/Riscani/Sector.txt")
        self.wsPt = self.wbPt.active
        # if not self.decControl:
        #     self.wsAng = self.wbAng.active

        self.ptList = []
        for i in range(2, self.wsPt.max_row + 1):
            self.ptList.append(self.wsPt.cell(row=i, column=1).value)
        myCompleter = QCompleter(self.ptList)
        myCompleter.setCaseSensitivity(Qt.CaseInsensitive)
        self.ptLine.setCompleter(myCompleter)

        if not self.decControl:
            if not self.deranjControl:
                # Completez forma pentru emitent
                self.emList = []
                for i in self.angajati.find():
                    if i["gr_ts"] == "5":
                        self.emList.append(i["name"] + " " + "gr. " + i["gr_ts"])
                myCompleter = QCompleter(self.emList)
                myCompleter.setCaseSensitivity(Qt.CaseInsensitive)
                self.emLine.setCompleter(myCompleter)
                # Completez forma pentru sef
                self.sefList = []
                # print(self.angajati)
                for i in self.angajati.find():
                    self.sefList.append(i["name"] + " " + "gr. " + i["gr_ts"])
                myCompleter = QCompleter(self.sefList)
                myCompleter.setCaseSensitivity(Qt.CaseInsensitive)
                self.sfLine.setCompleter(myCompleter)
                if not self.alControl:
                    self.memEchLine.setCompleter(myCompleter)

            if self.deranjControl:
                # Completez forma pentru sef
                self.sefList = []
                for i in self.angajati.find():
                    self.sefList.append(i["name"])
                myCompleter = QCompleter(self.sefList)
                myCompleter.setCaseSensitivity(Qt.CaseInsensitive)
                self.sfLine.setCompleter(myCompleter)
                self.sectorList = []
                #Incarc sectoarele de MongoDB
                for i in self.sector:
                    i = i.strip()
                    self.sectorList.append(i)
                self.sectCombo.clear()
                self.sectCombo.addItems(self.sectorList)
                self.sectCombo.setCurrentText("Alege:")
                self.f10kvList = []
                for i in self.f10kv:
                    i = i.strip()
                    self.f10kvList.append(i)
                myCompleter = QCompleter(self.f10kvList)
                myCompleter.setCaseSensitivity(Qt.CaseInsensitive)
                self.f10kvLine.setCompleter(myCompleter)

    def loadPtSec(self):
        if self.data.at[self.modRow, 0] == "UN":
            self.wbPt = load_workbook("Bundle/Ungheni/PT_Ungheni.xlsx")
        if self.data.at[self.modRow, 0] == "FL":
            self.wbPt = load_workbook("Bundle/Falesti/PT_Falesti.xlsx")
        if self.data.at[self.modRow, 0] == "GL":
            self.wbPt = load_workbook("Bundle/Glodeni/PT_Glodeni.xlsx")
        if self.data.at[self.modRow, 0] == "RS":
            self.wbPt = load_workbook("Bundle/Riscani/PT_Riscani.xlsx")
        self.wsPt = self.wbPt.active

    # Incarc informatia pentru sectoare
    def loadSect(self):
        if self.ofCombo.currentText() == "Ungheni":
            self.sector = open("Bundle/Ungheni/Sector.txt")
        if self.ofCombo.currentText() == "Falesti":
            self.sector = open("Bundle/Falesti/Sector.txt")
        if self.ofCombo.currentText() == "Glodeni":
            self.sector = open("Bundle/Glodeni/Sector.txt")
        if self.ofCombo.currentText() == "Riscani":
            self.sector = open("Bundle/Riscani/Sector.txt")

        self.sectorList = []
        if self.ofCombo.currentText() != "Toate":
            self.sectorList = ["Toate"]
            for i in self.sector:
                i = i.strip()
                self.sectorList.append(i)
        self.sectCombo.clear()
        self.sectCombo.addItems(self.sectorList)
        # self.sectCombo.setCurrentText("Alege:")

    def intTrig(self):
        if not self.passControl:
        #Formez dereastra dialogului
            self.dialInt = QDialog()
            self.dialInt.setWindowFlags(Qt.WindowCloseButtonHint)
            self.dialInt.setWindowIcon(QIcon('ataman_logo'))
            self.dialInt.setWindowTitle('Log In:')
            self.dialInt.setStyleSheet('background-color: #424242;')

        #Oficiul
            ofFrame = QFrame()
            ofFrame.setFrameShape(QFrame.StyledPanel)
            ofFrame.setStyleSheet("background-color: #315240; color: #e3e3e3")

            ofLabel = QLabel("Oficiul:")
        #Incarc oficiile din MongoDB
            self.loadOficii()
            self.ofCombo = QComboBox()
            self.ofCombo.setStyleSheet('background-color: #315240; color: #e3e3e3;  height:20')
            self.ofCombo.setFixedWidth(100)
            self.ofCombo.addItems(self.ofList)
            self.ofCombo.currentTextChanged.connect(self.loadAng)

            hbox = QHBoxLayout()
            hbox.addWidget(ofLabel)
            hbox.addWidget(self.ofCombo)
            ofFrame.setLayout(hbox)

        #User Frame
            uFrame = QFrame()
            uFrame.setFrameShape(QFrame.StyledPanel)
            uFrame.setStyleSheet("background-color: #315240; color: #e3e3e3")

            self.loadMongoUN()


            uLabel = QLabel("User:")


            self.uCombo = QComboBox()
            self.uCombo.setStyleSheet('background-color: #315240; color: #e3e3e3;  height:20')
            self.uCombo.setFixedWidth(100)
            self.uCombo.setEditable(True)
            # self.uCombo.setFont(QFont("Calibri", QFont.Bold))
            self.loadAng()
            # self.uCombo.setCurrentText("Alege:")


            hbox = QHBoxLayout()
            hbox.addWidget(uLabel)
            hbox.addWidget(self.uCombo)
            uFrame.setLayout(hbox)

        # User Frame
            psFrame = QFrame()
            psFrame.setFrameShape(QFrame.StyledPanel)
            psFrame.setStyleSheet("background-color: #315240; color: #e3e3e3")

            psLabel = QLabel("Parola:")
            self.psText = PasswordEdit()
            self.psText.setStyleSheet('background-color: #ffffff; color: #050505')
            self.psText.setFixedWidth(100)
        # self.psText.setEchoMode(QLineEdit.Password)

            hbox = QHBoxLayout()
            hbox.addWidget(psLabel)
            hbox.addWidget(self.psText)
            psFrame.setLayout(hbox)

        # Buttons Section (butoanele "Ok", "Cancel"
            btFrame = QFrame()
            btFrame.setFrameShape(QFrame.StyledPanel)

            btOk = QPushButton('Ok')
            btOk.setStyleSheet('color: #e3e3e3')
            btOk.clicked.connect(self.okPass)

            btCancel = QPushButton('Cancel')
            btCancel.setStyleSheet('color: #e3e3e3')
            btCancel.clicked.connect(self.cancelPass)

            hbox = QHBoxLayout()
            hbox.addWidget(btOk)
            hbox.addWidget(btCancel)
            btFrame.setLayout(hbox)

            vbox = QVBoxLayout()
            vbox.addWidget(ofFrame)
            vbox.addWidget(uFrame)
            vbox.addWidget(psFrame)
            vbox.addWidget(btFrame)
            self.dialInt.setLayout(vbox)

            self.dialInt.exec()


    def setTrig(self):
        # Dialog Window Create
        self.dialBox = QDialog()
        self.dialBox.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialBox.setWindowIcon(QIcon('ataman_logo'))
        self.dialBox.setWindowTitle('Setari')
        self.dialBox.setStyleSheet('background-color: #424242;')

        setGenFrame = QFrame()
        setGenFrame.setFrameShape(QFrame.StyledPanel)

        #Destinatia registru autorizatii
        setRegFrame = QFrame()
        setRegFrame.setFrameShape(QFrame.StyledPanel)
        setRegFrame.setStyleSheet('background-color: #5c3838;')

        setRegLabel = QLabel("Introduceti destinatia registru AUTORIZATII (excel):")
        setRegLabel.setStyleSheet('color: #e3e3e3;')
        self.setRegLine = QLineEdit()
        self.setRegLine.setFixedWidth(250)
        self.setRegLine.setStyleSheet('background-color: #e3e3e3;')
        rxDt = QRegExp("")
        myDtValidator = QRegExpValidator(rxDt)
        self.setRegLine.setValidator(myDtValidator)

        setRegButton = QPushButton()
        setRegButton.setIcon(QIcon("folder.ico"))
        setRegButton.setIconSize(QSize(20, 20))
        setRegButton.clicked.connect(self.destRegAl)

        hbox = QHBoxLayout()
        hbox.addWidget(setRegLabel)
        hbox.addWidget(self.setRegLine)
        hbox.addWidget(setRegButton)
        setRegFrame.setLayout(hbox)

        # Destinatia SAIDI
        setSaidiFrame = QFrame()
        setSaidiFrame.setFrameShape(QFrame.StyledPanel)
        setSaidiFrame.setStyleSheet('background-color: #5c3838;')

        setSaidiLabel = QLabel("Introduceti destinatia SAIDI (excel):")
        setSaidiLabel.setStyleSheet('color: #e3e3e3; margin-right: 70%')
        self.setSaidiLine = QLineEdit()
        self.setSaidiLine.setStyleSheet('background-color: #e3e3e3;')
        rxDt = QRegExp("")
        myDtValidator = QRegExpValidator(rxDt)
        self.setSaidiLine.setValidator(myDtValidator)

        setSaidiButton = QPushButton()
        setSaidiButton.setIcon(QIcon("folder.ico"))
        setSaidiButton.setIconSize(QSize(20, 20))
        setSaidiButton.clicked.connect(self.destSaidi)

        hbox = QHBoxLayout()
        hbox.addWidget(setSaidiLabel)
        hbox.addWidget(self.setSaidiLine)
        hbox.addWidget(setSaidiButton)
        setSaidiFrame.setLayout(hbox)

        # Destinatia RAPORT PDJT
        setRapFrame = QFrame()
        setRapFrame.setFrameShape(QFrame.StyledPanel)
        setRapFrame.setStyleSheet('background-color: #5c3838;')

        setRapLabel = QLabel("Introduceti destinatia RAPORT PDJT (excel):")
        setRapLabel.setStyleSheet('color: #e3e3e3; margin-right: 28%')
        self.setRapLine = QLineEdit()
        self.setRapLine.setStyleSheet('background-color: #e3e3e3;')
        rxDt = QRegExp("")
        myDtValidator = QRegExpValidator(rxDt)
        self.setRapLine.setValidator(myDtValidator)

        setRapButton = QPushButton()
        setRapButton.setIcon(QIcon("folder.ico"))
        setRapButton.setIconSize(QSize(20, 20))
        setRapButton.clicked.connect(self.destRap)

        hbox = QHBoxLayout()
        hbox.addWidget(setRapLabel)
        hbox.addWidget(self.setRapLine)
        hbox.addWidget(setRapButton)
        setRapFrame.setLayout(hbox)

        setBtFrame = QFrame()
        # setBtFrame.setFrameShape(QFrame.StyledPanel)

        setBtOk = QPushButton('Ok')
        setBtOk.setStyleSheet('color: #e3e3e3')
        setBtOk.setFixedWidth(100)
        setBtOk.clicked.connect(self.destOk)
        btCancel = QPushButton('Cancel')
        btCancel.setStyleSheet('color: #e3e3e3')
        btCancel.setFixedWidth(100)
        btCancel.clicked.connect(self.dialBox.close)

        setEmptyLb = QLabel()

        hbox = QHBoxLayout()
        hbox.addWidget(setEmptyLb)
        hbox.addWidget(setBtOk)
        hbox.addWidget(btCancel)
        setBtFrame.setLayout(hbox)

        vbox = QVBoxLayout()
        vbox.addWidget(setRegFrame)
        vbox.addWidget(setSaidiFrame)
        vbox.addWidget(setRapFrame)
        vbox.addWidget(setBtFrame)
        setGenFrame.setLayout(vbox)

        hbox = QHBoxLayout()
        hbox.addWidget(setGenFrame)
        self.dialBox.setLayout(hbox)

        self.destLoad()
        self.setRegLine.setText(self.wsDest.cell(row=1, column=2).value)
        self.setSaidiLine.setText(self.wsDest.cell(row=2, column=2).value)
        self.setRapLine.setText(self.wsDest.cell(row=3, column=2).value)

        self.dialBox.exec()

    def destRegAl(self):
        dlg = QFileDialog()
        myDirectory = dlg.getExistingDirectory()
        self.setRegLine.setText(myDirectory)
    def destSaidi(self):
        dlg = QFileDialog()
        myDirectory = dlg.getExistingDirectory()
        self.setSaidiLine.setText(myDirectory)
    def destRap(self):
        dlg = QFileDialog()
        myDirectory = dlg.getExistingDirectory()
        self.setRapLine.setText(myDirectory)

    def destOk(self):
        self.destLoad()
        if self.setRegLine.text() == "":
            self.msCall("destinatia pentru Registru AUTORIZATII excel!")
        elif self.setSaidiLine.text() == "":
            self.msCall("destinatia pentru SAIDI excel!")
        elif self.setRapLine.text() == "":
            self.msCall("destinatia pentru RAPORTUL PDJT excel!")
        else:
            self.wsDest.cell(row=1, column=2).value = self.setRegLine.text()
            self.wsDest.cell(row=2, column=2).value = self.setSaidiLine.text()
            self.wsDest.cell(row=3, column=2).value = self.setRapLine.text()
            try:
                self.wbDest.save(self.myDest)
            except PermissionError:
                self.msSecCall("Datele nu s-au introdus in fisierul Destination.xlsx!")
            self.dialBox.close()

    def destLoad(self):
        self.myDest = os.path.abspath(".") + "/Bundle/Destination.xlsx"
        self.wbDest = load_workbook(self.myDest)
        self.wsDest = self.wbDest["Destination"]

    def deranjTrig(self):
        #Verific ferestrele din meniul New, le folosesc la loadPT
        self.deranjControl = True
        self.alControl = False
        self.decControl = False

        self.dialBox = QDialog()
        self.dialBox.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialBox.setWindowIcon(QIcon('ataman_logo'))
        self.dialBox.setWindowTitle('Deranjament')
        self.dialBox.setStyleSheet('background-color: #424242;')

        # Oficiul
        ofFrame = QFrame()
        ofFrame.setFrameShape(QFrame.StyledPanel)
        ofFrame.setStyleSheet('background-color: #544637;')


        ofLabel = QLabel()
        ofLabel.setText('Oficiul:')
        ofLabel.setStyleSheet('color: #e3e3e3')

        self.ofCombo = QComboBox()
        self.ofCombo.setStyleSheet('background-color: #544637; color: #e3e3e3;  height:20')
        self.ofCombo.setFixedWidth(150)
        self.ofCombo.addItems(self.ofList)
        self.ofCombo.currentTextChanged.connect(self.loadPt)

        hbox = QHBoxLayout()
        hbox.addWidget(ofLabel)
        hbox.addWidget(self.ofCombo)
        ofFrame.setLayout(hbox)

        # Transmis
        transmisFrame= QFrame()
        transmisFrame.setFrameShape(QFrame.StyledPanel)
        transmisFrame.setStyleSheet('background-color: #544637;')

        transmisLabel = QLabel()
        transmisLabel.setText('Transmis:')
        transmisLabel.setStyleSheet('color: #e3e3e3')

        self.sfLine = QLineEdit()
        self.sfLine.setStyleSheet('background-color: #ffffff')
        self.sfLine.setFixedWidth(150)

        hbox = QHBoxLayout()
        hbox.addWidget(transmisLabel)
        hbox.addWidget(self.sfLine)
        transmisFrame.setLayout(hbox)

        # Instalatia
        instalatiaFrame = QFrame()
        instalatiaFrame.setFrameShape(QFrame.StyledPanel)
        instalatiaFrame.setStyleSheet('background-color: #544637;')

        instalatiaLabel = QLabel()
        instalatiaLabel.setText('Instalatia:')
        instalatiaLabel.setStyleSheet('color: #e3e3e3')

        instalatiaList = []
        instalatia = open("Bundle/Instalatia_deranj.txt", "r")
        for i in instalatia:
            i = i.strip()
            instalatiaList.append(i)

        self.instalatiaCombo = QComboBox()
        self.instalatiaCombo.setStyleSheet('background-color: #544637; color: #e3e3e3;  height:20')
        self.instalatiaCombo.setFixedWidth(150)
        self.instalatiaCombo.addItems(instalatiaList)
        self.instalatiaCombo.setEditable(True)
        self.instalatiaCombo.setCurrentText("Alege:")

        hbox = QHBoxLayout()
        hbox.addWidget(instalatiaLabel)
        hbox.addWidget(self.instalatiaCombo)
        instalatiaFrame.setLayout(hbox)

        # Sectorul
        sectFrame = QFrame()
        sectFrame.setFrameShape(QFrame.StyledPanel)
        sectFrame.setStyleSheet('background-color: #544637;')

        sectLabel = QLabel()
        sectLabel.setText('Sectorul:')
        sectLabel.setStyleSheet('color: #e3e3e3')

        self.sectCombo = QComboBox()
        self.sectCombo.setStyleSheet('background-color: #544637; color: #e3e3e3;  height:20')
        self.sectCombo.setFixedWidth(150)
        self.sectCombo.setEditable(True)

        hbox = QHBoxLayout()
        hbox.addWidget(sectLabel)
        hbox.addWidget(self.sectCombo)
        sectFrame.setLayout(hbox)

        # F 10kV
        f10kvFrame = QFrame()
        f10kvFrame.setFrameShape(QFrame.StyledPanel)
        f10kvFrame.setStyleSheet('background-color: #544637;')

        f10kvLabel = QLabel()
        f10kvLabel.setText('Fid.10kV nr:')
        f10kvLabel.setStyleSheet('color: #e3e3e3')

        self.f10kvLine = QLineEdit()
        self.f10kvLine.setStyleSheet('background-color: #ffffff')
        self.f10kvLine.setFixedWidth(150)

        hbox = QHBoxLayout()
        hbox.addWidget(f10kvLabel)
        hbox.addWidget(self.f10kvLine)
        f10kvFrame.setLayout(hbox)

        #PT frame
        ptFrame = QFrame()
        ptFrame.setFrameShape(QFrame.StyledPanel)
        ptFrame.setStyleSheet('background-color: #544637;')

        ptLabel = QLabel()
        ptLabel.setText('PT, Fider:')
        ptLabel.setStyleSheet('color: #e3e3e3')

        # PT Section
        self.ptLine = QLineEdit()
        self.ptLine.setStyleSheet('background-color: #ffffff')
        self.ptLine.setFixedWidth(150)
        self.loadPt()
        self.ptLine.setText("PT")
        rxPt = QRegExp("(PT|PD)\d.(dot)?.(dot)")
        rxPt.setCaseSensitivity(Qt.CaseInsensitive)
        myValidator = QRegExpValidator(rxPt)
        self.ptLine.setValidator(myValidator)

        # Pentru QlineEdit folosesc clasul MyLineEdit ca sa pot folosi FocusInEvent
        self.ptFidLine = QLineEdit()
        self.ptFidLine.setStyleSheet('background-color: #ffffff')
        self.ptFidLine.setFixedWidth(150)
        self.ptFidLine.setText("Fider nr.")
        rxGr = QRegExp("Fider nr.\d\d?\d?")
        myGrValidator = QRegExpValidator(rxGr)
        self.ptFidLine.setValidator(myGrValidator)

        gridPtFrame = QGridLayout()
        ptFrame.setLayout(gridPtFrame)

        gridPtFrame.addWidget(ptLabel, 0, 0)
        gridPtFrame.addWidget(self.ptLine, 0, 1)
        gridPtFrame.addWidget(self.ptFidLine, 1, 1)

        # Continutul
        continFrame = QFrame()
        continFrame.setFrameShape(QFrame.StyledPanel)
        continFrame.setStyleSheet('background-color: #544637;')

        continLabel = QLabel()
        continLabel.setText('Continutul:')
        continLabel.setStyleSheet('color: #e3e3e3')

        self.continText = QTextEdit()
        self.continText.setStyleSheet("Background-color: rgb(255, 255, 255)")
        self.continText.setFixedHeight(45)
        self.continText.setFixedWidth(150)

        hbox = QHBoxLayout()
        hbox.addWidget(continLabel)
        hbox.addWidget(self.continText)
        continFrame.setLayout(hbox)

        # Buttons Section (butoanele "Ok", "Cancel"
        btFrame = QFrame()
        btFrame.setFrameShape(QFrame.StyledPanel)

        btOk = QPushButton('Ok')
        btOk.setStyleSheet('color: #e3e3e3')
        btOk.clicked.connect(self.okDeranjTrigger)

        btCancel = QPushButton('Cancel')
        btCancel.setStyleSheet('color: #e3e3e3')
        btCancel.clicked.connect(self.dialBox.close)

        hbox = QHBoxLayout()
        hbox.addWidget(btOk)
        hbox.addWidget(btCancel)
        btFrame.setLayout(hbox)

        # Gneral grid
        grid = QGridLayout()
        self.dialBox.setLayout(grid)
        grid.addWidget(ofFrame, 1, 0)
        grid.addWidget(transmisFrame, 2, 0)
        grid.addWidget(sectFrame, 3, 0)
        grid.addWidget(instalatiaFrame, 4, 0)
        grid.addWidget(f10kvFrame, 5, 0)
        grid.addWidget(ptFrame, 6, 0)
        grid.addWidget(continFrame, 7, 0)
        grid.addWidget(btFrame, 8, 0)

        self.dialBox.exec()

    def decTrig(self):
        self.decControl = True
        self.deranjControl = False

        self.dialBox = QDialog()
        self.dialBox.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialBox.setWindowIcon(QIcon('ataman_logo'))
        self.dialBox.setWindowTitle('Deconectare neplanificata')
        self.dialBox.setStyleSheet('background-color: #424242;')

        # Nr. Deconectarii
        nrDecFrame = QFrame()
        nrDecFrame.setFrameShape(QFrame.StyledPanel)
        nrDecFrame.setStyleSheet('background-color: #5c3838;')


        ofLabel = QLabel()
        ofLabel.setText('Oficiul:')
        ofLabel.setStyleSheet('color: #e3e3e3')

    #Incarc oficiile din MongoDB
        # self.loadOficii()
        self.ofCombo = QComboBox()
        self.ofCombo.setStyleSheet('background-color: #5c3838; color: #e3e3e3;  height:20')
        self.ofCombo.setFixedWidth(140)
        self.ofCombo.addItems(self.ofList)
        self.ofCombo.currentTextChanged.connect(self.loadPt)

        hbox = QHBoxLayout()
        # hbox.addWidget(nrLabel)
        hbox.addWidget(ofLabel)
        hbox.addWidget(self.ofCombo)
        nrDecFrame.setLayout(hbox)

        #PT frame
        ptFrame = QFrame()
        ptFrame.setFrameShape(QFrame.StyledPanel)
        ptFrame.setStyleSheet('background-color: #5c3838;')

        ptLabel = QLabel()
        ptLabel.setText('Numarul PT, Fider:')
        ptLabel.setStyleSheet('color: #e3e3e3; margin-right: 25%')

        # PT Section
        self.ptLine = QLineEdit()
        self.ptLine.setStyleSheet('background-color: #ffffff')
        self.ptLine.setFixedWidth(140)
        self.loadPt()
        # self.ptList = []
        # for i in range(2, self.wsPt.max_row + 1):
        #     self.ptList.append(self.wsPt.cell(row=i, column=1).value)
        # ptCompleter = QCompleter(self.ptList)
        # ptCompleter.setCaseSensitivity(Qt.CaseInsensitive)
        # self.ptLine.setCompleter(ptCompleter)
        self.ptLine.setText("PT")
        rxPt = QRegExp("(PT|PD)\d.(dot)?.(dot)")
        rxPt.setCaseSensitivity(Qt.CaseInsensitive)
        myValidator = QRegExpValidator(rxPt)
        self.ptLine.setValidator(myValidator)

        # Pentru QlineEdit folosesc clasul MyLineEdit ca sa pot folosi FocusInEvent
        self.ptFidLine = QLineEdit()
        self.ptFidLine.setStyleSheet('background-color: #ffffff')
        self.ptFidLine.setFixedWidth(140)
        self.ptFidLine.setText("Fider nr.")
        rxGr = QRegExp("Fider nr.\d\d?\d?")
        myGrValidator = QRegExpValidator(rxGr)
        self.ptFidLine.setValidator(myGrValidator)

        gridPtFrame = QGridLayout()
        ptFrame.setLayout(gridPtFrame)

        gridPtFrame.addWidget(ptLabel, 0, 0)
        gridPtFrame.addWidget(self.ptLine, 0, 1)
        gridPtFrame.addWidget(self.ptFidLine, 1, 1)

        # Data frame
        dtFrame = QFrame()
        dtFrame.setFrameShape(QFrame.StyledPanel)
        dtFrame.setStyleSheet('background-color: #5c3838;')

        dtLabel = QLabel()
        dtLabel.setText('Data si ora deconectarii:')
        dtLabel.setStyleSheet('color: #e3e3e3')

        self.dtLine = DateLineEdit()
        self.dtLine.setStyleSheet("Background-color: rgb(255, 255, 255)")
        self.dtLine.setFixedWidth(140)

        rxDt = QRegExp("\d\d.\d\d.\d\d \d:\d")
        myDtValidator = QRegExpValidator(rxDt)
        self.dtLine.setValidator(myDtValidator)
        pyDateTime = datetime.datetime.now().strftime("%d.%m.%y %H:%M")
        self.dtLine.setText(pyDateTime)

        hbox = QHBoxLayout()
        hbox.addWidget(dtLabel)
        hbox.addWidget(self.dtLine)
        dtFrame.setLayout(hbox)

        cauzaFrame = QFrame()
        cauzaFrame.setFrameShape(QFrame.StyledPanel)
        cauzaFrame.setStyleSheet('background-color: #314652;')

        cauzaLabel = QLabel()
        cauzaLabel.setText('Cauza deconectarii:')
        cauzaLabel.setStyleSheet('color: #e3e3e3; margin-right: 15%')

        wbCauza = load_workbook(os.path.abspath(".") + "/Bundle/Cauza_deconectare.xlsx")
        wsCauza = wbCauza.active
        cauzaList = []
        for i in range(1, wsCauza.max_row + 1):
            cauzaList.append(wsCauza.cell(row=i, column=1).value)
        self.cauzaCombo = QComboBox()
        self.cauzaCombo.setStyleSheet('background-color: #314652; color: #e3e3e3;  height:20')
        self.cauzaCombo.setFixedWidth(140)
        self.cauzaCombo.setEditable(True)
        self.cauzaCombo.addItems(cauzaList)
        # self.cauzaCombo.setFixedWidth(120)

        hbox = QHBoxLayout()
        hbox.addWidget(cauzaLabel)
        hbox.addWidget(self.cauzaCombo)
        cauzaFrame.setLayout(hbox)

        termenFrame = QFrame()
        termenFrame.setFrameShape(QFrame.StyledPanel)
        termenFrame.setStyleSheet('background-color: #314652;')

        termenLabel = QLabel()
        termenLabel.setText('Termen reglementat:')
        termenLabel.setStyleSheet('color: #e3e3e3')

        termenList = ['Incadrat', 'Neincadrat']
        self.termenCombo = QComboBox()
        self.termenCombo.setStyleSheet('background-color: #314652; color: #e3e3e3;  height:20')
        self.termenCombo.setFixedWidth(140)
        self.termenCombo.addItems(termenList)

        hbox = QHBoxLayout()
        hbox.addWidget(termenLabel)
        hbox.addWidget(self.termenCombo)
        termenFrame.setLayout(hbox)

        # Buttons Section (butoanele "Ok", "Cancel"
        btFrame = QFrame()
        btFrame.setFrameShape(QFrame.StyledPanel)

        btOk = QPushButton('Ok')
        btOk.setStyleSheet('color: #e3e3e3')
        btOk.clicked.connect(self.okDecTrigger)

        btCancel = QPushButton('Cancel')
        btCancel.setStyleSheet('color: #e3e3e3')
        btCancel.clicked.connect(self.dialBox.close)

        hbox = QHBoxLayout()
        hbox.addWidget(btOk)
        hbox.addWidget(btCancel)
        btFrame.setLayout(hbox)

        # Gneral grid
        grid = QGridLayout()
        self.dialBox.setLayout(grid)
        grid.addWidget(nrDecFrame, 1, 0)
        grid.addWidget(ptFrame, 2, 0)
        grid.addWidget(dtFrame, 3, 0)
        grid.addWidget(cauzaFrame, 4, 0)
        grid.addWidget(termenFrame, 5, 0)
        grid.addWidget(btFrame, 6, 0)

        self.dialBox.exec()

    def dsTrig(self):
        self.alControl = False
        self.decControl = False
        self.deranjControl = False

        # Incarc registru dupa data din Excel
        self.dtRegAl()

        # Dialog Window Create
        self.dialBox = QDialog()
        self.dialBox.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialBox.setWindowIcon(QIcon('ataman_logo'))
        self.dialBox.setWindowTitle('Dispozitie')
        self.dialBox.setStyleSheet('background-color: #424242;')

        #Nr. Dispozitiei frame
        nrDsFrame = QFrame()
        nrDsFrame.setFrameShape(QFrame.StyledPanel)
        nrDsFrame.setStyleSheet('background-color: #314652;')

        for i in range(self.wsRegAl.max_row, 3, -1):
            if self.wsRegAl.cell(row=i, column=3).value != None:
                self.nrAlDs = int(self.wsRegAl.cell(row=i, column=3).value) + 1
                break

        nrLabel = QLabel()
        nrLabel.setStyleSheet('color: #e3e3e3;')
        nrLabel.setText('Numarul dispozitiei: ' + str(self.nrAlDs))

        ofLabel = QLabel()
        ofLabel.setText('Oficiul:')
        ofLabel.setStyleSheet('color: #e3e3e3; padding-left:10px')

        # Incarc oficiile din MongoDB
        # self.loadOficii()
        self.ofCombo = QComboBox()
        self.ofCombo.setStyleSheet('background-color: #314652; color: #e3e3e3;  height:20')
        self.ofCombo.setFixedWidth(100)
        self.ofCombo.addItems(self.ofList)
        self.ofCombo.currentTextChanged.connect(self.loadPt)

        hbox = QHBoxLayout()
        hbox.addWidget(nrLabel)
        hbox.addWidget(ofLabel)
        hbox.addWidget(self.ofCombo)
        nrDsFrame.setLayout(hbox)

        # Sef de lucrari Frame (Sef de lucrari, grupa TS), Load Excel files
        sefFrame = QFrame()
        sefFrame.setFrameShape(QFrame.StyledPanel)
        sefFrame.setStyleSheet('background-color: #314652;')

        sefLabel = QLabel()
        sefLabel.setStyleSheet('color: #e3e3e3;')
        sefLabel.setText('Sef de lucrari:')

        self.sfLine = QLineEdit()
        self.sfLine.setStyleSheet('background-color: #ffffff')
        self.sfLine.setFixedWidth(180)

        hbox = QHBoxLayout()
        hbox.addWidget(sefLabel)
        hbox.addWidget(self.sfLine)
        sefFrame.setLayout(hbox)

        # Membrii echipei frame
        memEchFrame = QFrame()
        memEchFrame.setFrameShape(QFrame.StyledPanel)
        memEchFrame.setStyleSheet('background-color: #314652;')

        memEchLabel = QLabel()
        memEchLabel.setStyleSheet('color: #e3e3e3;')
        memEchLabel.setText('Membrii echipei:')

        self.memEchLine = QLineEdit()
        self.memEchLine.setStyleSheet('background-color: #ffffff')
        self.memEchLine.setFixedWidth(180)

        hbox = QHBoxLayout()
        hbox.addWidget(memEchLabel)
        hbox.addWidget(self.memEchLine)
        memEchFrame.setLayout(hbox)

        # Emitent Frame
        emFrame = QFrame()
        emFrame.setFrameShape(QFrame.StyledPanel)
        emFrame.setStyleSheet('background-color: #314652;')

        emLabel = QLabel()
        emLabel.setStyleSheet('color: #e3e3e3;')
        emLabel.setText('Emitent :')

        self.emLine = QLineEdit()
        self.emLine.setStyleSheet('background-color: #ffffff')
        self.emLine.setFixedWidth(180)

        hbox = QHBoxLayout()
        hbox.addWidget(emLabel)
        hbox.addWidget(self.emLine)
        emFrame.setLayout(hbox)

        # Se executa frame
        exFrame = QFrame()
        exFrame.setFrameShape(QFrame.StyledPanel)
        exFrame.setStyleSheet('background-color: #6b6843;')

        exLabel = QLabel()
        exLabel.setStyleSheet('color: #e3e3e3;')
        exLabel.setText('Se executa:')

        # Instalatia section
        wb = load_workbook('Bundle/Instalatia.xlsx')
        ws = wb.active
        instList = []
        for i in range(1, ws.max_row + 1):
            instList.append(ws.cell(row=i, column=1).value)

        # Widgetul QlineEdit este inlocuit cu clasul MyLineEdit creat
        # mai sus pentru a putea folosi FocusIN si OUT Event
        self.instLine = QComboBox()
        self.instLine.addItems(instList)
        self.instLine.setEditable(True)
        self.instLine.setStyleSheet('color: #e3e3e3;  height:20')
        self.instLine.setFixedWidth(180)

        # PT Sectiom
        self.ptLine = QLineEdit()
        self.ptLine.setStyleSheet('background-color: #ffffff')
        self.loadPt()
        self.ptLine.setText("PT")
        self.ptLine.setFixedWidth(180)

        # Pentru QlineEdit folosesc clasul MyLineEdit ca sa pot folosi FocusInEvent
        self.ptFidLine = QLineEdit()
        self.ptFidLine.setStyleSheet('background-color: #ffffff')
        # self.ptFidLine.setFixedWidth(20)
        self.ptFidLine.setText("Fider nr.")
        rxGr = QRegExp("Fider nr.\d(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)"
                       "(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)")
        myGrValidator = QRegExpValidator(rxGr)
        self.ptFidLine.setValidator(myGrValidator)
        self.ptFidLine.setFixedWidth(180)

        # Lucrari Section
        wb = load_workbook('Bundle/Lucrarile.xlsx')
        ws = wb.active
        lucrList = []
        for i in range(1, ws.max_row + 1):
            lucrList.append(ws.cell(row=i, column=1).value)
        self.lucrLine = QComboBox()
        self.lucrLine.setStyleSheet('color: #e3e3e3;  height:20')
        self.lucrLine.addItems(lucrList)
        self.lucrLine.setEditable(True)
        self.lucrLine.setCurrentText("Lucrarile efectuate:")
        self.lucrLine.setFixedWidth(180)

        gridExFrame = QGridLayout()
        exFrame.setLayout(gridExFrame)

        gridExFrame.addWidget(exLabel, 0, 0)
        gridExFrame.addWidget(self.instLine, 0, 1)
        gridExFrame.addWidget(self.ptLine, 1, 1)
        gridExFrame.addWidget(self.ptFidLine, 2, 1)
        gridExFrame.addWidget(self.lucrLine, 3, 1)
        # gridExFrame.addWidget(self.decCombo, 4, 1)

        # Buttons Section (butoanele "Ok", "Cancel"
        btFrame = QFrame()
        btFrame.setFrameShape(QFrame.StyledPanel)

        btOk = QPushButton('Ok')
        btOk.setStyleSheet('color: #e3e3e3')
        btOk.clicked.connect(self.okDsTrigger)

        btCancel = QPushButton('Cancel')
        btCancel.setStyleSheet('color: #e3e3e3')
        btCancel.clicked.connect(self.dialBox.close)

        hbox = QHBoxLayout()
        hbox.addWidget(btOk)
        hbox.addWidget(btCancel)
        btFrame.setLayout(hbox)


        # Gneral grid
        grid = QGridLayout()
        self.dialBox.setLayout(grid)
        grid.addWidget(nrDsFrame, 1, 0)
        grid.addWidget(sefFrame, 2, 0)
        grid.addWidget(memEchFrame, 3, 0)
        grid.addWidget(emFrame, 4, 0)
        grid.addWidget(exFrame, 5, 0)
        grid.addWidget(btFrame, 6, 0)

        self.dialBox.exec()

    def alTrig(self):
        self.alControl = True
        self.decControl = False
        self.deranjControl = False

        #Incarc registru dupa data din Excel
        self.dtRegAl()

        # Dialog Window Create
        self.dialBox = QDialog()
        self.dialBox.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialBox.setWindowIcon(QIcon('ataman_logo'))
        self.dialBox.setWindowTitle('Autorizatie')
        self.dialBox.setStyleSheet('background-color: #424242;')

        # Dialog First Frame (Nr. autorizatie, Oficiul)
        nrautFrame = QFrame()
        nrautFrame.setFrameShape(QFrame.StyledPanel)
        nrautFrame.setStyleSheet('background-color: #315240;')

        #Incarc registrul din excel
        for i in range(self.wsRegAl.max_row, 3, -1):
            if self.wsRegAl.cell(row=i, column=2).value != None:
                self.nrAlDs = int(self.wsRegAl.cell(row=i, column=2).value) + 1
                break

        nrLabel = QLabel()
        nrLabel.setStyleSheet('color: #e3e3e3;')
        nrLabel.setText('Numarul autorizatiei: ' + str(self.nrAlDs))

        ofLabel = QLabel()
        ofLabel.setText('Oficiul:')
        ofLabel.setStyleSheet('color: #e3e3e3; padding-left:10px')

        # Incarc oficiile din MongoDB
        # self.loadOficii()
        self.ofCombo = QComboBox()
        self.ofCombo.setStyleSheet('background-color: #315240; color: #e3e3e3;  height:20')
        self.ofCombo.setFixedWidth(80)
        self.ofCombo.addItems(self.ofList)
        self.ofCombo.currentTextChanged.connect(self.loadPt)

        hbox = QHBoxLayout()
        hbox.addWidget(nrLabel)
        hbox.addWidget(ofLabel)
        hbox.addWidget(self.ofCombo)
        nrautFrame.setLayout(hbox)

        # Sef de lucrari Frame (Sef de lucrari, grupa TS), Load Excel files
        sefFrame = QFrame()
        sefFrame.setFrameShape(QFrame.StyledPanel)
        sefFrame.setStyleSheet('background-color: #315240;')

        sefLabel = QLabel()
        sefLabel.setStyleSheet('color: #e3e3e3;')
        sefLabel.setText('Sef de lucrari:')

        self.sfLine = QLineEdit()
        self.sfLine.setStyleSheet('background-color: #ffffff')
        self.sfLine.setFixedWidth(180)

        gridSefFrame = QGridLayout()
        sefFrame.setLayout(gridSefFrame)

        gridSefFrame.addWidget(sefLabel, 0, 0)
        gridSefFrame.addWidget(self.sfLine, 0, 1)

        # Emitent Frame
        emFrame = QFrame()
        emFrame.setFrameShape(QFrame.StyledPanel)
        emFrame.setStyleSheet('background-color: #315240;')

        emLabel = QLabel()
        emLabel.setStyleSheet('color: #e3e3e3;')
        emLabel.setText('Emitent :        ')

        self.emLine = QLineEdit()
        self.emLine.setStyleSheet('background-color: #ffffff')
        self.emLine.setFixedWidth(180)

        gridEmFrame = QGridLayout()
        emFrame.setLayout(gridEmFrame)

        gridEmFrame.addWidget(emLabel, 0, 0)
        gridEmFrame.addWidget(self.emLine, 0, 1)

        # Se executa frame
        exFrame = QFrame()
        exFrame.setFrameShape(QFrame.StyledPanel)
        exFrame.setStyleSheet('background-color: #6b6843;')

        exLabel = QLabel()
        exLabel.setStyleSheet('color: #e3e3e3; margin-right: 2')
        exLabel.setText('Se executa:')

        # Instalatia section
        wb = load_workbook('Bundle/Instalatia.xlsx')
        ws = wb.active
        instList = []
        for i in range(1, ws.max_row + 1):
            instList.append(ws.cell(row=i, column=1).value)

        # Widgetul QlineEdit este inlocuit cu clasul MyLineEdit creat
        # mai sus pentru a putea folosi FocusIN si OUT Event
        self.instLine = QComboBox()
        self.instLine.addItems(instList)
        self.instLine.setEditable(True)
        self.instLine.setStyleSheet('color: #e3e3e3;  height:20')
        self.instLine.setFixedWidth(180)

        # PT Section
        self.ptLine = QLineEdit()
        self.ptLine.setStyleSheet('background-color: #ffffff')
        self.loadPt()

        self.ptLine.setText("PT")
        rxPt = QRegExp("(PT|PD)\d.(dot)?.(dot)")
        rxPt.setCaseSensitivity(Qt.CaseInsensitive)
        myValidator = QRegExpValidator(rxPt)
        self.ptLine.setValidator(myValidator)
        self.ptLine.setFixedWidth(180)

        #Pentru QlineEdit folosesc clasul MyLineEdit ca sa pot folosi FocusInEvent
        self.ptFidLine = QLineEdit()
        self.ptFidLine.setStyleSheet('background-color: #ffffff')
        self.ptFidLine.setFixedWidth(180)
        self.ptFidLine.setText("Fider nr.")
        rxGr = QRegExp("Fider nr.\d\d?\d?")
        myGrValidator = QRegExpValidator(rxGr)
        self.ptFidLine.setValidator(myGrValidator)

        #Lucrari Section
        wb = load_workbook('Bundle/Lucrarile.xlsx')
        ws = wb.active
        lucrList = []
        for i in range(1, ws.max_row + 1):
            lucrList.append(ws.cell(row=i, column=1).value)
        self.lucrLine = QComboBox()
        self.lucrLine.setStyleSheet('color: #e3e3e3;  height:20')
        self.lucrLine.addItems(lucrList)
        self.lucrLine.setEditable(True)
        self.lucrLine.setCurrentText("Lucrarile efectuate:")
        self.lucrLine.setFixedWidth(180)

        self.smNr = QLineEdit()
        self.smNr.setStyleSheet('background-color: #ffffff')
        self.smNr.setText("SM nr.")
        rxGr = QRegExp("SM nr.\d(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)"
                       "(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)(\d?|,)")
        myGrValidator = QRegExpValidator(rxGr)
        self.smNr.setValidator(myGrValidator)
        self.smNr.setFixedWidth(180)

        self.decCombo = QComboBox()
        self.decCombo.setStyleSheet('background-color: #6b6843; color: #e3e3e3; height:20')
        self.decCombo.addItem("Programat")
        self.decCombo.addItem("Neprogramat")
        self.decCombo.addItem("Fara deconectari")


        gridExFrame = QGridLayout()
        exFrame.setLayout(gridExFrame)

        gridExFrame.addWidget(exLabel, 0, 0)
        gridExFrame.addWidget(self.instLine, 0, 1)
        gridExFrame.addWidget(self.ptLine, 1, 1)
        gridExFrame.addWidget(self.ptFidLine, 2, 1)
        gridExFrame.addWidget(self.lucrLine, 3, 1)
        gridExFrame.addWidget(self.smNr, 4, 1)
        gridExFrame.addWidget(self.decCombo, 5, 1)

        # Buttons Section (butoanele "Ok", "Cancel"
        btFrame = QFrame()
        btFrame.setFrameShape(QFrame.StyledPanel)

        btOk = QPushButton('Ok')
        btOk.setStyleSheet('color: #e3e3e3')
        btOk.clicked.connect(self.okTrigger)

        btCancel = QPushButton('Cancel')
        btCancel.setStyleSheet('color: #e3e3e3')
        btCancel.clicked.connect(self.dialBox.close)

        gridBt = QGridLayout()
        btFrame.setLayout(gridBt)
        gridBt.addWidget(btOk, 0, 0)
        gridBt.addWidget(btCancel, 0, 1)

        # General grid
        grid = QGridLayout()
        self.dialBox.setLayout(grid)
        grid.addWidget(nrautFrame,1,0)
        grid.addWidget(sefFrame, 2, 0)
        grid.addWidget(emFrame, 3, 0)
        grid.addWidget(exFrame, 4,0)
        grid.addWidget(btFrame, 5, 0)

        self.dialBox.exec()

    #  Functie prelucrare erori cimpul autorizatii
    def msCall(self, intrCimp):
        self.intrCimp = intrCimp
        self.intrMB = QMessageBox()
        self.intrMB.setIcon(QMessageBox.Warning)
        self.intrMB.setWindowTitle('Atentie!')
        self.intrMB.setText('Introduceti datele in cimpul:\n' + self.intrCimp)
        self.intrMB.setWindowIcon(QIcon('ataman_logo.ico'))
        self.intrMB.exec()

    def msSecCall(self, myMess):
        self.myMess = myMess
        self.secMB = QMessageBox()
        self.secMB.setIcon(QMessageBox.Information)
        self.secMB.setWindowTitle('Pentru informare:')
        self.secMB.setText(self.myMess)
        self.secMB.setWindowIcon(QIcon('ataman_logo.ico'))
        self.secMB.exec()

    #Functie populez Biroul dispecerului cu Autorizatie
    def centrAlPop(self):

        # Incarc registru dupa data din Excel
        self.dtRegAl()
        self.deranjControlPop = False

        if self.tabWindowControl == True:
            self.tabWindow.close()
            # self.tabWindowControl = False
        self.centralReg = []
        for i in range(self.wsRegAl.max_row, 4, -1):
            myColumn = []
            for j in range(1, self.wsRegAl.max_column + 1):
                # data = np.array(centralReg.append(self.wsRegAl.cell(row=i, column=j)))
                myColumn.append(self.wsRegAl.cell(row=i, column=j).value)
            self.centralReg.append(myColumn)
        self.data = pd.DataFrame(self.centralReg)
        header = ["Oficiul", "Nr. \nAL", "Nr. \nDS", "Instalatia", \
                  "PT", "Localitatea", "Nr. \nFider", "Lucrarile efectuate", \
                  "Sef de lucrari\nsau supraveghetor\n(numele, prenumele)\ngrupa de securitate)", \
                  "Membrii formatiei\nce lucreaza pe DS\n(numele, prenumele,\n grupa de securitate)",
                  "Lucratorul care a emis\n autorizatia (dispoziție)\n(numele, prenumele,\ngrupa de securitate)",\
                  "Cu deconectare",\
                  "Masurile tehnice\n de asigurare a securitatii \nlucratorilor, cu indicarea \ndeconectarilor necesare,\nlocurilor de montare\na legaturilor la pamint",\
                  "Semnaturile lucratorilor\ncare au executat \ninstruirea periodica \nsi care au fost instruiti",\
                  "Ora de admitere", "Ora terminarii\n lucrarilor", "Persoana care\n a inregistrat AL, DS \n(numele, prenumele),\n data",
                  "Persoana care\na admis echipa\n(numele, prenumele)"]
        self.table = QTableView()
        self.model = TableModel(self.data, header)

        self.table.setModel(self.model)
        self.table.setWordWrap(True)
        self.table.setTextElideMode(Qt.ElideMiddle)
        self.table.resizeColumnsToContents()
        # self.table.setColumnWidth(1, 10)
        # self.table.setColumnWidth(2, 10)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 100)
        self.table.setColumnWidth(6, 50)
        self.table.setColumnWidth(7, 150)
        self.table.setColumnWidth(9, 130)
        self.table.setColumnWidth(12, 150)
        self.table.setColumnWidth(13, 150)
        self.table.setColumnWidth(14, 100)
        self.table.setColumnWidth(15, 100)
        self.table.setColumnWidth(16, 120)
        self.table.setColumnWidth(17, 120)
        self.table.resizeRowsToContents()
        self.table.setStyleSheet("Background-color: rgb(200, 200, 200)")
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().hide()

        regFrame = QFrame()
        regAlTitle = QLabel()

        regAlTitle.setText("Registru de autorizatii Of.:")
        regAlTitle.setStyleSheet("padding-left: 50%; font-size:24px; color:rgb(191, 60, 60)")
        # regAlTitle.move(0, 100)
        # self.loadOficii()
        ofList = ['Toate']
        ofList = ofList + self.ofList
        self.ofCombo = QComboBox()
        self.ofCombo.addItems(ofList)
        # self.ofComboReg.setStyleSheet("margin-right:1400%")
        self.ofCombo.setStyleSheet("padding-left:10%; font-size:12px")
        self.ofCombo.setFixedHeight(25)
        self.ofCombo.setFixedWidth(100)
        self.ofCombo.currentTextChanged.connect(self.ofChangeReg)

        emptyLb = QLabel("")
        emptyLb2 = QLabel("")
        emptyLb3 = QLabel("")
        emptyLb4 = QLabel("")
        # emptyLb.setStyleSheet("margin-right:1400%")

        hbox = QHBoxLayout()
        hbox.addWidget(regAlTitle)
        hbox.addWidget(self.ofCombo)
        hbox.addWidget(emptyLb)
        hbox.addWidget(emptyLb2)
        hbox.addWidget(emptyLb3)
        hbox.addWidget(emptyLb4)
        hbox.setStretch(1, 1)

        vbox = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addWidget(self.table)
        regFrame.setLayout(vbox)

        self.tabWindow = QMdiSubWindow()
        self.myMidi.addSubWindow(self.tabWindow)
        self.tabWindow.setWindowIcon(QIcon(QPixmap(1, 1)))
        self.tabWindow.setWidget(regFrame)
        self.tabWindow.setGeometry(100, 100, 1000, 600)
        self.tabWindow.showMaximized()
        self.tabWindowControl = True
        self.tabWindow.show()

        self.table.doubleClicked.connect(self.showAl)

    def changeExec(self):
        for i in range(len(self.data.count(axis=1))):
            if self.execCheck.isChecked() == True:
                if self.table.isRowHidden(i) == False:
                    if self.data.at[i, 11] != "Neexecutat":
                        self.table.hideRow(i)
            elif self.execCheck.isChecked() == False:
                self.table.showRow(i)
                if self.ofCombo.currentText() == "Toate":
                    pass
                elif self.ofCombo.currentText() != "Toate":
                    if self.data.at[i, 0] != self.ofVar:
                        self.table.hideRow(i)
                    if self.sectCombo.currentText() != "Toate":
                        if self.data.at[i, 3] != self.sectCombo.currentText():
                            self.table.hideRow(i)
        # print(self.execCheck.isChecked())

    def changeSect(self):
        self.execCheck.setChecked(False)
        for i in range(len(self.data.count(axis=1))):
            self.table.showRow(i)
            if self.ofCombo.currentText() != "Toate":
                if self.data.at[i, 0] != self.ofVar:
                    self.table.hideRow(i)
            if self.sectCombo.currentText() != "Toate":
                if self.data.at[i, 3] != self.sectCombo.currentText():
                    self.table.hideRow(i)


    def ofChangeReg(self):
        self.abrOficii()
        for i in range(len(self.data.count(axis=1))):
            self.table.showRow(i)
            if self.ofCombo.currentText() != "Toate":
                if self.data.at[i, 0] != self.ofVar:
                    self.table.hideRow(i)
        if self.deranjControlPop == True:
            self.execCheck.setChecked(False)
            self.loadSect()

    def showAl(self):
        self.modRow = self.table.currentIndex().row()
        # self.modColumn = self.table.currentIndex().column()
        newRegEx = self.wsRegAl.max_row - self.modRow
        if self.wsRegAlLink.cell(row=newRegEx, column=1).value:
            webbrowser.open(self.wsRegAlLink.cell(row=newRegEx, column=1).value)
        else:
            self.msSecCall("Nu exista link!")

    def decTabWindow(self):
        if self.tabDecControl == True:
            self.tabWindDec.close()
            # self.tabDecControl = False

        self.dtContrDecZl()

        self.centralReg = []
        for i in range(7, self.wsDecPT.max_row + 1):
            myColumn = []
            for j in range(1, self.wsDecPT.max_column - 1):
                # data = np.array(centralReg.append(self.wsRegAl.cell(row=i, column=j)))
                myColumn.append(self.wsDecPT.cell(row=i, column=j).value)
            self.centralReg.append(myColumn)
        self.dataDec = pd.DataFrame(self.centralReg)
        # print(self.data)
        # pd.set_option('display.max_columns', None)
        header = ["Nr.", "Oficiul", "PT", "Fider", "Data si ora deconectarii", "Data si ora conectarii", "Durata intreruperii",
                  "Consumatori casnici", "Consumatori non-casnici",
                  "Total",
                  "Localitate", "Cauza deconectarii",
                  "Termenul\nreglementat\n6/12"]
        self.tableDecPt = QTableView()
        self.model = TableModel(self.dataDec, header)

        self.tableDecPt.setModel(self.model)
            # self.table.setWordWrap(True)
        self.tableDecPt.setTextElideMode(Qt.ElideMiddle)
        self.tableDecPt.resizeColumnsToContents()
        # self.table.columnsWidth(9, 300)
        self.tableDecPt.resizeRowsToContents()
        self.tableDecPt.setStyleSheet("Background-color: rgb(200, 200, 200)")
        self.tableDecPt.setSelectionBehavior(QAbstractItemView.SelectRows)
        for i in range(0, 13):
            self.tableDecPt.setColumnWidth(i, 140)
        self.tableDecPt.setColumnWidth(4, 160)
        self.tableDecPt.setColumnWidth(5, 160)
        self.tableDecPt.verticalHeader().hide()

        #Creez a II-lea tabel
        self.centralReg = []
        for i in range(5, 9):
            myColumn = []
            for j in range(1, 12):
                # data = np.array(centralReg.append(self.wsRegAl.cell(row=i, column=j)))
                if isinstance(self.wsNrSol.cell(row=i, column=j).value, int):
                    myColumn.append(str(self.wsNrSol.cell(row=i, column=j).value))
                else:
                    myColumn.append(self.wsNrSol.cell(row=i, column=j).value)
            self.centralReg.append(myColumn)
        self.dataDec = pd.DataFrame(self.centralReg)

        header = ["Nr.", "Oficiul", "Total", "Remediat", "In Executare",
                  "Termen reglementat\n6/12 nerespectat",
                  "Depasire, ore", "Cauza SA RED-Nord",
                  "Cauza IS ME", "Cauza consumator",
                  "Altele"]
        # verHeader = ["UN", "FL", "GL", "RS"]

        self.tableNrSol = QTableView()
        self.model = TableModelII(self.dataDec, header)

        self.tableNrSol.setModel(self.model)
        # self.table.setWordWrap(True)
        self.tableNrSol.setTextElideMode(Qt.ElideMiddle)
        self.tableNrSol.resizeColumnsToContents()
        # self.tableNrSol.setColumnWidth(9, 300)
        self.tableNrSol.resizeRowsToContents()
        self.tableNrSol.setStyleSheet("Background-color: rgb(200, 200, 200)")
        for i in range(0, 11):
            self.tableNrSol.setColumnWidth(i, 140)
        self.tableNrSol.verticalHeader().hide()
        # self.tableNrSol.clicked.connect(self.nrSolChange)

        # Creez a III-lea tabel
        self.centralReg = []
        for i in range(13, 17):
            myColumn = []
            for j in range(1, 12):
                # data = np.array(centralReg.append(self.wsRegAl.cell(row=i, column=j)))
                if isinstance(self.wsNrSol.cell(row=i, column=j).value, int):
                    myColumn.append(str(self.wsNrSol.cell(row=i, column=j).value))
                else:
                    myColumn.append(self.wsNrSol.cell(row=i, column=j).value)
            self.centralReg.append(myColumn)
        self.dataDec = pd.DataFrame(self.centralReg)


        # header = []
        self.tableInf = QTableView()
        self.model = TableModelIII(self.dataDec)

        self.tableInf.setModel(self.model)
        # self.table.setWordWrap(True)
        self.tableInf.setTextElideMode(Qt.ElideMiddle)
        self.tableInf.resizeColumnsToContents()
        # self.tableInf.setColumnWidth(9, 300)
        self.tableInf.resizeRowsToContents()
        self.tableInf.setStyleSheet("Background-color: rgb(200, 200, 200)")
        # self.tableInf.horizontalHeader().hide()
        self.tableInf.verticalHeader().hide()
        for i in range(3, 11):
            self.tableInf.setColumnHidden(i, True)
        self.tableInf.setColumnWidth(2, 1460)


        #I Frame Deconectari PT

        rapDecFrame = QFrame()

        rapDecTitle = QLabel("Raport deconectari zilnice")
        rapDecTitle.setStyleSheet("padding-left: 50%; font-size:24px; color:rgb(191, 60, 60)")

        rapData = QLabel(" " + datetime.datetime.now().strftime("%d.%m.%y"))
        rapData.setStyleSheet("font-size:24px")

        mySpace = QLabel(" ")

        hbox = QHBoxLayout()
        hbox.addWidget(rapDecTitle)
        hbox.addWidget(rapData)
        hbox.addWidget(mySpace)
        hbox.addWidget(mySpace)
        hbox.addWidget(mySpace)
        hbox.addWidget(mySpace)

        decPT = QLabel("Deconectari PT")
        decPT.setStyleSheet("padding-left: 100%; font-size:14px; color:rgb(0, 0, 0)")

        vbox = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addWidget(decPT)
        vbox.addWidget(self.tableDecPt)
        # vbox.addWidget(spaceFrame)
        rapDecFrame.setLayout(vbox)

        #Numar de solicitari, informatii suplimentare

        nrSolFrame = QFrame()

        nrSol = QLabel("Numar solicitari")
        nrSol.setStyleSheet("padding-left: 100%; font-size:14px; color:rgb(0, 0, 0)")

        infSupl = QLabel("Informatii suplimentare")
        infSupl.setStyleSheet("padding-left: 100%; font-size:14px; color:rgb(0, 0, 0)")

        vbox = QVBoxLayout()
        vbox.addWidget(nrSol)
        vbox.addWidget(self.tableNrSol)
        vbox.addWidget(infSupl)
        vbox.addWidget(self.tableInf)
        nrSolFrame.setLayout(vbox)

        myScrRes = app.primaryScreen()
        myScrAvailable = myScrRes.availableGeometry()
        myHeight = round(myScrAvailable.height()/4)

        mySplitter = QSplitter(Qt.Vertical)
        mySplitter.addWidget(rapDecFrame)
        mySplitter.addWidget(nrSolFrame)
        mySplitter.setSizes([myHeight, 1])

        mySplitter.handle(1).setStyleSheet("Background-color: rgb(191, 60, 60)")
        # mySplitter.setHandleWidth(10)
        # print(mySplitter.handleWidth())

        self.tabWindDec = QMdiSubWindow()
        self.myMidi.addSubWindow(self.tabWindDec)
        self.tabWindDec.setWindowIcon(QIcon(QPixmap(1, 1)))
        self.tabWindDec.setWidget(mySplitter)
        self.tabWindDec.setGeometry(100, 100, 1000, 600)
        self.tabWindDec.showMaximized()
        self.tabDecControl = True
        self.tabWindDec.show()

    def executat(self):
        pyDateTime = datetime.datetime.now()
        self.data.at[self.modRow, self.modColumn] = \
            "Executat: " + self.uCombo.currentText() + "\n" + pyDateTime.strftime("%d.%m.%y")
        self.db.deranjamente.update_one({
            "nr_ordine": self.data.at[self.modRow, 1]
        }, {
            "$set": {
                "starea": self.data.at[self.modRow, self.modColumn]
            }
        })

    def intrerupere(self):
        # self.tabWindow.close()
        # self.centrAlPop()
        self.data.at[self.modRow, self.modColumn] = \
            "Intrerupere"

        # Fac update la Excel dupa introducerea datelor la "Nou inregistrata"
        newRegEx = self.wsRegAl.max_row - self.modRow
        self.wsRegAl.cell(row=newRegEx, column=16).value = \
            self.data.at[self.modRow, self.modColumn]
        try:
            self.wbRegAl.save(self.regFile)
        except PermissionError:
            self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                           "(nu este permisa introducerea datelor). Incercati mai tirziu!")
            self.centrAlPop()

    def semneaza(self):
        if self.data.at[self.modRow, self.modColumn] == "-//-":
            pyDateTime = datetime.datetime.now()
            self.data.at[self.modRow, self.modColumn] = self.uCombo.currentText() +\
                "\n" + pyDateTime.strftime("%d.%m.%y")

            if self.tabWindowControl == True:
                newRegEx = self.wsRegAl.max_row - self.modRow
                self.wsRegAl.cell(row=newRegEx, column=14).value = \
                    self.data.at[self.modRow, self.modColumn]
                try:
                    self.wbRegAl.save(self.regFile)
                except PermissionError:
                    self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                                "(nu este permisa introducerea datelor). Incercati mai tirziu!")
                    self.centrAlPop()

            if self.deranjControlPop:
                self.db.deranjamente.update_one({
                     "nr_ordine": self.data.at[self.modRow, 1]
                    }, {
                        "$set": {
                            "responsabil": self.data.at[self.modRow, self.modColumn]
                        }
                    })


    def admEch(self):
        if self.data.at[self.modRow, self.modColumn] == "Nou inregistrata":
            # myDateTime = QDateTime.currentDateTime()
            pyDateTime = datetime.datetime.now()
            self.data.at[self.modRow, self.modColumn] = \
                pyDateTime.strftime("%d.%m.%Y %H:%M")
            self.data.at[self.modRow, self.modColumn+1] = \
                "/ ... /"
            self.data.at[self.modRow, self.modColumn + 3] = \
                self.uCombo.currentText() + "\n" + "admis"
                # pyDateTime.strftime("%d")+"."+pyDateTime.strftime("%m")+\
                # "."+pyDateTime.strftime("%Y")+" "+pyDateTime.strftime("%H")+\
                # ":"+pyDateTime.strftime("%M")


            #Fac update la Excel dupa introducerea datelor la "Nou inregistrata"
            newRegEx = self.wsRegAl.max_row - self.modRow
            self.wsRegAl.cell(row=newRegEx, column=15).value = \
                self.data.at[self.modRow, self.modColumn]
            self.wsRegAl.cell(row=newRegEx, column=16).value = \
                self.data.at[self.modRow, self.modColumn+1]
            self.wsRegAl.cell(row=newRegEx, column=18).value = self.uCombo.currentText() + \
                                                               "\n" + "admis"
            self.wsRegAl.cell(row=newRegEx, column=18).alignment = \
                Alignment(horizontal="center", vertical="center")
            try:
                self.wbRegAl.save(self.regFile)
            except PermissionError:
                self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                             "(nu este permisa introducerea datelor). Incercati mai tirziu!")
                self.centrAlPop()

    # Functie chemata din context menu pentru neefectuarea lucrarilor
    def nuLucr(self):
        self.data.at[self.modRow, self.modColumn] = "Nu s-a lucrat"
        self.data.at[self.modRow, self.modColumn+1] = "/ --- /"

        # Fac update la Excel dupa introducerea datelor la "Nou inregistrata"
        newRegEx = self.wsRegAl.max_row - self.modRow
        self.wsRegAl.cell(row=newRegEx, column=15).value = \
            self.data.at[self.modRow, self.modColumn]
        self.wsRegAl.cell(row=newRegEx, column=16).value = \
            self.data.at[self.modRow, self.modColumn+1]
        try:
            self.wbRegAl.save(self.regFile)
        except PermissionError:
            self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                           "(nu este permisa introducerea datelor). Incercati mai tirziu!")
            self.centrAlPop()

    #Functie terminarea lucrarilor
    def termLucr(self):
        pyDateTime = datetime.datetime.now()
        self.data.at[self.modRow, self.modColumn] = \
            pyDateTime.strftime("%d.%m.%y %H:%M")
        valueMin = self.data.at[self.modRow, self.modColumn - 1]
        strToDate = datetime.datetime.strptime(valueMin, "%d.%m.%Y %H:%M")
        # print(strToDate)
        self.data.at[self.modRow, self.modColumn - 1] = \
            strToDate.strftime("%d.%m.%y %H:%M")

        # Fac update la Excel dupa introducerea datelor la "/ ... /"
        newRegEx = self.wsRegAl.max_row - self.modRow
        self.wsRegAl.cell(row=newRegEx, column=15).value = \
            self.data.at[self.modRow, self.modColumn - 1]
        self.wsRegAl.cell(row=newRegEx, column=16).value = \
            self.data.at[self.modRow, self.modColumn]
        try:
            self.wbRegAl.save(self.regFile)
        except PermissionError:
            self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                           "(nu este permisa introducerea datelor). Incercati mai tirziu!")
            self.centrAlPop()
            self.erContrAl = True
        # Introduc datele in Excel deconectari
        #Controlez daca exista mapa cu anul
        self.dtContrSaidi()
        self.dtAnAnual()
        self.loadPtSec()

        if self.data.at[self.modRow, 11] == "Programat" and not self.erContrAl:
            myMaxRow = self.wsDecProg.max_row + 1
            self.wsDecProg.cell(row=myMaxRow, column=1).value = \
                self.wsDecProg.cell(row=myMaxRow - 1, column=1).value + 1
            self.wsDecProg.cell(row=myMaxRow, column=1).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=2).value = \
                self.data.at[self.modRow, 0]
            self.wsDecProg.cell(row=myMaxRow, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=3).value = \
                datetime.date.today().strftime("%d.%m.%y")
            self.wsDecProg.cell(row=myMaxRow, column=3).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=4).value = \
                "JT"
            self.wsDecProg.cell(row=myMaxRow, column=4).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=5).value = \
                self.data.at[self.modRow, 4]
            self.wsDecProg.cell(row=myMaxRow, column=5).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=6).value = \
                self.data.at[self.modRow, 6]
            self.wsDecProg.cell(row=myMaxRow, column=6).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=7).value = \
                0
            self.wsDecProg.cell(row=myMaxRow, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=8).value = \
                1
            self.wsDecProg.cell(row=myMaxRow, column=8).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=9).value = \
                1
            self.wsDecProg.cell(row=myMaxRow, column=9).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=10).value = \
                self.data.at[self.modRow, self.modColumn - 1]
            self.wsDecProg.cell(row=myMaxRow, column=10).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecProg.cell(row=myMaxRow, column=11).value = \
                self.data.at[self.modRow, self.modColumn]
            self.wsDecProg.cell(row=myMaxRow, column=11).alignment = \
                Alignment(horizontal="center", vertical="center")

            #Calculez si pun orele in coloana 12
            valueMaxDate_11 = self.wsDecProg.cell(row=myMaxRow, column=11).value
            valueMaxDate_10 = self.wsDecProg.cell(row=myMaxRow, column=10).value
            strToDate_11 = datetime.datetime.strptime(valueMaxDate_11, "%d.%m.%y %H:%M")
            strToDate_10 = datetime.datetime.strptime(valueMaxDate_10, "%d.%m.%y %H:%M")
            delta_11_10 = strToDate_11 - strToDate_10
            myList = str(delta_11_10).split()
            try:
                if int(myList[0]):
                    myDateSplit = myList[2].split(":")
                    myDayPlus = 24 * int(myList[0]) + int(myDateSplit[0])
                    myDateStr = str(myDayPlus) + ":" + myDateSplit[1] + ":" + myDateSplit[2]
                    myDeltaHour = myDateStr
            except ValueError:
                myDeltaHour = str(delta_11_10)
            self.wsDecProg.cell(row=myMaxRow, column=12).value = myDeltaHour
            self.wsDecProg.cell(row=myMaxRow, column=12).alignment = \
                Alignment(horizontal="center", vertical="center")

            #Calculez si pun numarul de consumatori si localitatea
            for i in range(2, self.wsPt.max_row + 1):
                if self.wsPt.cell(row=i, column=1).value == self.data.at[self.modRow, 4]:
                    self.totNrCas = int(self.wsPt.cell(row=i, column=4).value)
                    self.fidNrCas = round(self.totNrCas/3)
                    if self.fidNrCas > 65:
                        self.fidNrCas = random.randrange(60, 70)
                    self.wsDecProg.cell(row=myMaxRow, column=13).value = self.fidNrCas
                    self.wsDecProg.cell(row=myMaxRow, column=13).alignment = \
                        Alignment(horizontal="center", vertical="center")

                    self.totNrEc = self.wsPt.cell(row=i, column=5).value
                    if self.totNrEc <= 2 and self.totNrCas == 0:
                        self.fidNrEc = 1
                    elif self.totNrEc > 12:
                        self.fidNrEc = random.randrange(2, 5)
                    else:
                        self.fidNrEc = round(self.totNrEc/3)
                    self.wsDecProg.cell(row=myMaxRow, column=14).value = self.fidNrEc
                    self.wsDecProg.cell(row=myMaxRow, column=14).alignment = \
                        Alignment(horizontal="center", vertical="center")
                    self.wsDecProg.cell(row=myMaxRow, column=16).value = self.wsPt.cell(row=i, column=2).value

            self.wsDecProg.cell(row=myMaxRow, column=15).value = \
                str(self.wsDecProg.cell(row=myMaxRow, column=13).value + \
                self.wsDecProg.cell(row=myMaxRow, column=14).value)
            self.wsDecProg.cell(row=myMaxRow, column=15).alignment = \
                Alignment(horizontal="center", vertical="center")

            self.wsDecProg.cell(row=myMaxRow, column=19).value = \
                self.data.at[self.modRow, 7]

            #Introduc datele in Excel analiza anuala
            anAnualMaxRow = self.wsAnAnualP.max_row + 1
            self.wsAnAnualP.cell(row=anAnualMaxRow, column=1).value = \
                self.data.at[self.modRow, 0]
            self.wsAnAnualP.cell(row=anAnualMaxRow, column=2).value = \
                self.data.at[self.modRow, 4] + self.data.at[self.modRow, 6]
            self.wsAnAnualP.cell(row=anAnualMaxRow, column=3).value = \
                str(self.wsDecProg.cell(row=myMaxRow, column=13).value + \
                    self.wsDecProg.cell(row=myMaxRow, column=14).value)
            self.wsAnAnualP.cell(row=anAnualMaxRow, column=4).value = \
                str(delta_11_10)
            try:
                self.wbAnAnual.save(self.fileAnAnual)
            except PermissionError:
                self.msSecCall("Datele din autorizatie, sectiunea PROGRAMAT \n"
                               "nu vor participa la analiza anuala (undeva este deschisa analiza anuala excel)!")
            try:
                self.wbDec.save(self.saidiFile)
            except PermissionError:
                self.msSecCall("Datele din autorizatie, sectiunea PROGRAMAT \n"
                               "nu s-au introdus in fisierul SAIDI excel (cineva foloseste aplicatia)!")

        if self.data.at[self.modRow, 11] == "Neprogramat" and not self.erContrAl:
            myMaxRow = self.wsDecNeProg.max_row + 1
            self.wsDecNeProg.cell(row=myMaxRow, column=1).value = \
                self.wsDecNeProg.cell(row=myMaxRow - 1, column=1).value + 1
            self.wsDecNeProg.cell(row=myMaxRow, column=1).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=2).value = \
                self.data.at[self.modRow, 0]
            self.wsDecNeProg.cell(row=myMaxRow, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=3).value = \
                datetime.date.today().strftime("%d.%m.%y")
            self.wsDecNeProg.cell(row=myMaxRow, column=3).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=4).value = \
                self.data.at[self.modRow, 4]
            self.wsDecNeProg.cell(row=myMaxRow, column=4).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=5).value = \
                self.data.at[self.modRow, 6]
            self.wsDecNeProg.cell(row=myMaxRow, column=5).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=6).value = \
                0
            self.wsDecNeProg.cell(row=myMaxRow, column=6).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=7).value = \
                0
            self.wsDecNeProg.cell(row=myMaxRow, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=8).value = \
                1
            self.wsDecNeProg.cell(row=myMaxRow, column=8).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=9).value = \
                1
            self.wsDecNeProg.cell(row=myMaxRow, column=9).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=10).value = \
                self.data.at[self.modRow, self.modColumn - 1]
            self.wsDecNeProg.cell(row=myMaxRow, column=10).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecNeProg.cell(row=myMaxRow, column=11).value = \
                self.data.at[self.modRow, self.modColumn]
            self.wsDecNeProg.cell(row=myMaxRow, column=11).alignment = \
                Alignment(horizontal="center", vertical="center")

            # Calculez si pun orele diferenta lor
            valueMaxDate_11 = self.wsDecNeProg.cell(row=myMaxRow, column=11).value
            valueMaxDate_10 = self.wsDecNeProg.cell(row=myMaxRow, column=10).value
            strToDate_11 = datetime.datetime.strptime(valueMaxDate_11, "%d.%m.%y %H:%M")
            strToDate_10 = datetime.datetime.strptime(valueMaxDate_10, "%d.%m.%y %H:%M")
            delta_11_10 = strToDate_11 - strToDate_10
            myList = str(delta_11_10).split()
            try:
                if int(myList[0]):
                    myDateSplit = myList[2].split(":")
                    myDayPlus = 24 * int(myList[0]) + int(myDateSplit[0])
                    myDateStr = str(myDayPlus) + ":" + myDateSplit[1] + ":" + myDateSplit[2]
                    myDeltaHour = myDateStr
            except ValueError:
                myDeltaHour = str(delta_11_10)
            self.wsDecNeProg.cell(row=myMaxRow, column=19).value = myDeltaHour
            self.wsDecNeProg.cell(row=myMaxRow, column=19).alignment = \
                Alignment(horizontal="center", vertical="center")

            # Calculez si pun numarul de consumatori si localitatea
            for i in range(2, self.wsPt.max_row + 1):
                if self.wsPt.cell(row=i, column=1).value == self.data.at[self.modRow, 4]:
                    self.totNrCas = int(self.wsPt.cell(row=i, column=4).value)
                    self.fidNrCas = round(self.totNrCas / 3)
                    if self.fidNrCas > 65:
                        self.fidNrCas = random.randrange(60, 70)
                    self.wsDecNeProg.cell(row=myMaxRow, column=25).value = self.fidNrCas
                    self.wsDecNeProg.cell(row=myMaxRow, column=25).alignment = \
                        Alignment(horizontal="center", vertical="center")

                    self.totNrEc = self.wsPt.cell(row=i, column=5).value
                    if self.totNrEc <= 2 and self.totNrCas == 0:
                        self.fidNrEc = 1
                    elif self.totNrEc > 12:
                        self.fidNrEc = random.randrange(2, 5)
                    else:
                        self.fidNrEc = round(self.totNrEc/3)
                    self.wsDecNeProg.cell(row=myMaxRow, column=26).value = self.fidNrEc
                    self.wsDecNeProg.cell(row=myMaxRow, column=26).alignment = \
                        Alignment(horizontal="center", vertical="center")
                    self.wsDecNeProg.cell(row=myMaxRow, column=27).value = \
                        self.wsDecNeProg.cell(row=myMaxRow, column=25).value + \
                        self.wsDecNeProg.cell(row=myMaxRow, column=26).value
                    self.wsDecNeProg.cell(row=myMaxRow, column=27).alignment = \
                        Alignment(horizontal="center", vertical="center")
                    self.wsDecNeProg.cell(row=myMaxRow, column=28).value = self.wsPt.cell(row=i, column=2).value
                    myLocalitate = self.wsPt.cell(row=i, column=2).value

            self.wsDecNeProg.cell(row=myMaxRow, column=32).value = \
                self.data.at[self.modRow, 7]
            try:
                self.wbDec.save(self.saidiFile)
            except PermissionError:
                self.msSecCall("Datele din autorizatie, sectiunea NEPROGRAMAT \n"
                               "nu s-au introdus in fisierul SAIDI excel (cineva foloseste aplicatia)!")
            self.dtContrDecZl()
            myMaxRow = self.wsDecPT.max_row + 1

            # if self.wsDecPT.cell(row=2, column=8).value == None:
            #     self.wsDecPT.cell(row=2, column=8).value = datetime.datetime.now().strftime("%d.%m.%Y")
            self.wsDecPT.cell(row=myMaxRow, column=1).value = \
                str(int(self.wsDecPT.cell(row=myMaxRow - 1, column=1).value) + 1)
            self.wsDecPT.cell(row=myMaxRow, column=1).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=2).value = \
                self.data.at[self.modRow, 0]
            self.wsDecPT.cell(row=myMaxRow, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=3).value = \
                self.data.at[self.modRow, 4]
            self.wsDecPT.cell(row=myMaxRow, column=4).value = \
                self.data.at[self.modRow, 6]
            self.wsDecPT.cell(row=myMaxRow, column=4).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=5).value = \
                self.data.at[self.modRow, self.modColumn - 1]
            self.wsDecPT.cell(row=myMaxRow, column=5).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=6).value = \
                self.data.at[self.modRow, self.modColumn]
            self.wsDecPT.cell(row=myMaxRow, column=6).alignment = \
                Alignment(horizontal="center", vertical="center")
            # Calculez si pun orele diferenta lor
            valueDate_6 = self.wsDecPT.cell(row=myMaxRow, column=6).value
            valueDate_5 = self.wsDecPT.cell(row=myMaxRow, column=5).value
            strToDate_6 = datetime.datetime.strptime(valueDate_6, "%d.%m.%y %H:%M")
            strToDate_5 = datetime.datetime.strptime(valueDate_5, "%d.%m.%y %H:%M")
            delta_6_5 = strToDate_6 - strToDate_5
            self.wsDecPT.cell(row=myMaxRow, column=7).value = str(delta_6_5)
            self.wsDecPT.cell(row=myMaxRow, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")

            # # Calculez si pun numarul de consumatori si localitatea
            # for i in range(2, self.wsPt.max_row + 1):
            #     if self.wsPt.cell(row=i, column=1).value == self.wsDecPT.cell(row=myMaxRow, column=3).value:
            #         self.totNrCas = int(self.wsPt.cell(row=i, column=4).value)
            #         self.fazaNrCas = round(self.totNrCas / 9)
            #         if self.fazaNrCas > 25:
            #             self.fazaNrCas = random.randrange(20, 30)
            self.wsDecPT.cell(row=myMaxRow, column=8).value = str(self.fidNrCas)
            self.wsDecPT.cell(row=myMaxRow, column=8).alignment = \
                    Alignment(horizontal="center", vertical="center")

                    # self.totNrEc = self.wsPt.cell(row=i, column=5).value
                    # if self.totNrEc <= 2 and self.totNrCas == 0:
                    #     self.fazaNrEc = 1
                    # elif self.totNrEc > 12:
                    #     self.fazaNrEc = random.randrange(1, 3)
                    # else:
                    #     self.fazaNrEc = round(self.totNrEc / 9)
            self.wsDecPT.cell(row=myMaxRow, column=9).value = str(self.fidNrEc)
            self.wsDecPT.cell(row=myMaxRow, column=9).alignment = \
                    Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=11).value = \
                    myLocalitate

            self.wsDecPT.cell(row=myMaxRow, column=10).value = \
                str(int(self.wsDecPT.cell(row=myMaxRow, column=8).value) + \
                    int(self.wsDecPT.cell(row=myMaxRow, column=9).value))
            self.wsDecPT.cell(row=myMaxRow, column=10).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsDecPT.cell(row=myMaxRow, column=12).value = self.data.at[self.modRow, 7]
            self.wsDecPT.cell(row=myMaxRow, column=13).value = "Incadrat"
            self.wsDecPT.cell(row=myMaxRow, column=13).alignment = \
                Alignment(horizontal="center", vertical="center")

            try:
                self.wbDecZil.save(self.rapFile)
            except PermissionError:
                self.msSecCall("Datele din autorizatie, sectiunea NEPROGRAMAT \n"
                               "nu s-au introdus in fisierul RAPORT PDJT excel (cineva foloseste aplicatia)!")

            # Introduc datele in Excel analiza anuala
            anAnualMaxRow = self.wsAnAnualN.max_row + 1
            self.wsAnAnualN.cell(row=anAnualMaxRow, column=1).value = \
                self.data.at[self.modRow, 0]
            self.wsAnAnualN.cell(row=anAnualMaxRow, column=2).value = \
                self.data.at[self.modRow, 4] + self.data.at[self.modRow, 6]
            self.wsAnAnualN.cell(row=anAnualMaxRow, column=3).value = \
                self.wsDecNeProg.cell(row=self.wsDecNeProg.max_row, column=25).value + \
                self.wsDecNeProg.cell(row=self.wsDecNeProg.max_row, column=26).value
            self.wsAnAnualN.cell(row=anAnualMaxRow, column=4).value = \
                str(delta_11_10)
            try:
                self.wbAnAnual.save(self.fileAnAnual)
            except PermissionError:
                self.msSecCall("Datele din autorizatie, sectiunea NEPROGRAMAT \n"
                               "nu vor participa la analiza anuala (undeva este deschisa analiza anuala excel)!")
        self.erContrAl = False

    #Functie eroare, sterg rindul
    def erFunc(self):
        #Sterg rindul in excel
        newRegEx = self.wsRegAl.max_row - self.modRow
        self.wsRegAl.delete_rows(newRegEx)
        try:
            self.wbRegAl.save(self.regFile)
        except PermissionError:
            self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                           "(nu este permisa stergerea datelor). Incercati mai tirziu!")
        self.tabWindow.close()
        self.centrAlPop()

    #Functie pentru context menu, admite echipa, folosesc try...except
    #ca sa depasesc eroarea cind nu-i tabelul
    def contextMenuEvent(self, event):
        try:
            if self.table.hasFocus():
                self.modRow = self.table.currentIndex().row()
                self.modColumn = self.table.currentIndex().column()

                if self.data.at[self.modRow, self.modColumn] == "Nou inregistrata" \
                        and self.data.at[self.modRow, self.modColumn-1] != "Nou inregistrata":
                    ctxMenu = QMenu(self)
                    admAction = ctxMenu.addAction("Echipa admisa")
                    nuLucrAction = ctxMenu.addAction("Nu s-a lucrat")
                    erAction = ctxMenu.addAction("Eroare")
                    ctxAction = ctxMenu.exec_(self.mapToGlobal(event.pos()))
                    if ctxAction == admAction:
                        self.admEch()
                    if ctxAction == nuLucrAction:
                        self.nuLucr()
                    if ctxAction == erAction:
                        self.erFunc()
                if self.data.at[self.modRow, self.modColumn] == "/ ... /":
                    ctxMenu = QMenu(self)
                    termAction = ctxMenu.addAction("Lucrarile s-au terminat")
                    intrAction = ctxMenu.addAction("Intrerupere")
                    ctxAction = ctxMenu.exec_(self.mapToGlobal(event.pos()))
                    if ctxAction == termAction:
                        self.termLucr()
                    if ctxAction == intrAction:
                        self.intrerupere()
                if self.data.at[self.modRow, self.modColumn] == "Intrerupere":
                    ctxMenu = QMenu(self)
                    termAction = ctxMenu.addAction("Lucrarile s-au terminat")
                    ctxAction = ctxMenu.exec_(self.mapToGlobal(event.pos()))
                    if ctxAction == termAction:
                        self.termLucr()
                if self.data.at[self.modRow, self.modColumn] == "-//-":
                    ctxMenu = QMenu(self)
                    termAction = ctxMenu.addAction("Semneaza!")
                    ctxAction = ctxMenu.exec_(self.mapToGlobal(event.pos()))
                    if ctxAction == termAction:
                        self.semneaza()
                if self.data.at[self.modRow, self.modColumn] == "Neexecutat":
                    ctxMenu = QMenu(self)
                    execAction = ctxMenu.addAction("Executat")
                    ctxAction = ctxMenu.exec_(self.mapToGlobal(event.pos()))
                    if ctxAction == execAction:
                        self.executat()
        except AttributeError:
            pass


    #  Functie pentru popularea Registrului de autorizatii
    def regPop(self):
        self.abrOficii()

        #Populez registru Excel
        myMaxRow = self.wsRegAl.max_row + 1
        self.wsRegAl.cell(row=myMaxRow, column=1).value = self.ofVar
        self.wsRegAl.cell(row=myMaxRow, column=1).alignment = \
            Alignment(horizontal="center", vertical="center")
        if self.okControl:
            self.wsRegAl.cell(row=myMaxRow, column=3).value = str(self.nrAlDs)
            self.wsRegAl.cell(row=myMaxRow, column=3).alignment = \
                Alignment(horizontal="center", vertical="center")
            self.wsRegAl.cell(row=myMaxRow, column=4).value = self.instLine.currentText()
            self.wsRegAl.cell(row=myMaxRow, column=4).alignment = \
                    Alignment(vertical="center")
            if self.ptLine.text() == "PT":
                self.wsRegAl.cell(row=myMaxRow, column=5).value = ""
            else:
                self.wsRegAl.cell(row=myMaxRow, column=5).value = self.ptLine.text()
                self.wsRegAl.cell(row=myMaxRow, column=5).alignment = \
                    Alignment(vertical="center")
            if self.ptFidLine.text() == "Fider nr.":
                self.wsRegAl.cell(row=myMaxRow, column=7).value = ""
            else:
                self.wsRegAl.cell(row=myMaxRow, column=7).value = self.ptFidLine.text()
                self.wsRegAl.cell(row=myMaxRow, column=7).alignment = \
                    Alignment(horizontal="center", vertical="center")

        else:
            self.wsRegAl.cell(row=myMaxRow, column=2).value = str(self.nrAlDs)
            self.wsRegAl.cell(row=myMaxRow, column=2).alignment = \
                Alignment(horizontal="center", vertical="center")
        if not self.okControl:
            self.wsRegAl.cell(row=myMaxRow, column=4).value = self.instLine.currentText()
            self.wsRegAl.cell(row=myMaxRow, column=4).alignment = \
                Alignment(vertical="center")
            self.wsRegAl.cell(row=myMaxRow, column=5).value = self.ptLine.text()
            self.wsRegAl.cell(row=myMaxRow, column=5).alignment = \
                Alignment(vertical="center")

            # Calculez localitatile cel 6
        for i in range(2, self.wsPt.max_row + 1):
            if self.wsPt.cell(row=i, column=1).value == self.ptLine.text():
                self.wsRegAl.cell(row=myMaxRow, column=6).value = self.wsPt.cell(row=i, column=2).value
        self.wsRegAl.cell(row=myMaxRow, column=6).alignment = \
            Alignment(vertical="center")
        if not self.okControl:
            self.wsRegAl.cell(row=myMaxRow, column=7).value = self.ptFidLine.text()
            self.wsRegAl.cell(row=myMaxRow, column=7).alignment = \
                Alignment(horizontal="center", vertical="center")
        self.wsRegAl.cell(row=myMaxRow, column=8).value = self.lucrLine.currentText()
        self.wsRegAl.cell(row=myMaxRow, column=8).alignment = \
            Alignment(vertical="center")
        self.wsRegAl.cell(row=myMaxRow, column=9).value = self.sfLine.text()
        self.wsRegAl.cell(row=myMaxRow, column=9).alignment = \
            Alignment(vertical="center")
        if self.okControl:
            if self.memEchLine.text() != "":
                self.wsRegAl.cell(row=myMaxRow, column=10).value = "Formatia: " + \
                    self.memEchLine.text()
                self.wsRegAl.cell(row=myMaxRow, column=10).alignment = \
                    Alignment(vertical="center")
        self.wsRegAl.cell(row=myMaxRow, column=11).value = self.emLine.text()
        self.wsRegAl.cell(row=myMaxRow, column=11).alignment = \
            Alignment(vertical="center")
        if self.okControl == False:
            self.wsRegAl.cell(row=myMaxRow, column=12).value = self.decCombo.currentText()
            self.wsRegAl.cell(row=myMaxRow, column=12).alignment = \
                Alignment(horizontal="center", vertical="center")
            if self.decCombo.currentText() == "Programat" or self.decCombo.currentText() == "Neprogramat":
                self.wsRegAl.cell(row=myMaxRow, column=13).value = "La " + self.ptLine.text() + \
                    "\nse deconect. " + self.ptFidLine.text() + ",\nse monteaza " + self.smNr.text()
                self.wsRegAl.cell(row=myMaxRow, column=13).alignment = \
                    Alignment(horizontal="center", vertical="center", wrapText="True")
        else:
            self.wsRegAl.cell(row=myMaxRow, column=12).value = "Fara deconectari"
            self.wsRegAl.cell(row=myMaxRow, column=12).alignment = \
                Alignment(horizontal="center", vertical="center")
        self.wsRegAl.cell(row=myMaxRow, column=14).value = "-//-"
        self.wsRegAl.cell(row=myMaxRow, column=14).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsRegAl.cell(row=myMaxRow, column=15).value = "Nou inregistrata"
        self.wsRegAl.cell(row=myMaxRow, column=15).alignment = \
            Alignment(horizontal="center", vertical="center")
        # self.wsRegAl.cell(row=myMaxRow, column=13).font = Font(name="Calibri", size=16)
        #adaug 3 spatii ora terminarii lucrarii
        self.wsRegAl.cell(row=myMaxRow, column=16).value = "Nou inregistrata"
        self.wsRegAl.cell(row=myMaxRow, column=16).alignment = \
            Alignment(horizontal="center", vertical="center")
        pyDateTime = datetime.datetime.now()
        self.wsRegAl.cell(row=myMaxRow, column=17).value = self.uCombo.currentText() + \
            "\n" + pyDateTime.strftime("%d.%m.%y")
        self.wsRegAl.cell(row=myMaxRow, column=17).alignment = \
            Alignment(horizontal="center", vertical="center")

        #Deschid fereastra cu autorizatii, creez linkul in Excel in Sheetul alLink
        dlg = QFileDialog()
        fileName = dlg.getOpenFileName()
        self.wsRegAlLink.cell(row=myMaxRow, column=1).value = fileName[0]

        try:
            self.wbRegAl.save(self.regFile)
        except PermissionError:
            self.msSecCall("Registru AUTORIZATIILOR excel este deschis undeva,\n"
                           "(nu este permisa introducerea datelor). Incercati mai tirziu!")

        self.centrAlPop()
        self.dialBox.close()

    def dtContrDecZl(self):
        self.destLoad()
        destPath = self.wsDest.cell(row=3, column=2).value
        # Controlez daca exista mapa cu anul, luna, daca nu exista o creez
        self.myYear = datetime.datetime.now().strftime("%Y")
        self.myMonth = datetime.datetime.now().strftime("%m")
        self.myDay = datetime.datetime.now().strftime("%d")
        self.myHour = datetime.datetime.now().strftime("%H")

        myDate = datetime.datetime.now()
        myDelta = datetime.timedelta(days=1)
        myDateNext = myDate + myDelta
        myDeltaMinus = datetime.timedelta(days=-1)
        myDateMinus = myDate + myDeltaMinus
        myDayMinus = myDateMinus.strftime("%d")

        myPath = destPath + "/" + str(self.myYear)
        dirControl = os.path.isdir(myPath)
        if not dirControl:
            os.mkdir(myPath)

        myMonthPath = myPath + "/" + str(self.myMonth)
        dirControl = os.path.isdir(myMonthPath)
        if not dirControl:
            os.mkdir(myMonthPath)

        if int(self.myHour) >= 8:
            self.rapFile = myMonthPath + "/" + str(self.myDay) + " Raport.xlsx"
        else:
            self.rapFile = myMonthPath + "/" + str(myDayMinus) + " Raport.xlsx"
        fileControl = os.path.isfile(self.rapFile)

        if not fileControl:
            myOriginal = os.path.abspath(".") + "/Bundle/Ungheni/Raport/Raport.xlsx"
            shutil.copyfile(myOriginal, self.rapFile)

        self.wbDecZil = load_workbook(self.rapFile)
        self.wsDecPT = self.wbDecZil["Deconectari PT"]
        self.wsNrSol = self.wbDecZil["Numar solicitari"]

        if int(self.myHour) >= 8:
            if self.wsDecPT.cell(row=2, column=8).value == None:
                self.wsDecPT.cell(row=2, column=8).value = myDate.strftime("%d.%m.%Y")
                self.wsDecPT.cell(row=2, column=9).value = "- " + myDateNext.strftime("%d.%m.%Y")
            if self.wsNrSol.cell(row=2, column=6).value == None:
                self.wsNrSol.cell(row=2, column=6).value = myDate.strftime("%d.%m.%Y")
                self.wsNrSol.cell(row=2, column=7).value = "- " + myDateNext.strftime("%d.%m.%Y")
                self.wbDecZil.save(self.rapFile)


    def dtContrSaidi(self):
        self.destLoad()
        destPath = self.wsDest.cell(row=2, column=2).value

        self.myYear = datetime.datetime.now().strftime("%Y")
        self.myMonth = datetime.datetime.now().strftime("%m")

        # Introduc datele in SAIDI cu controlul mapelor si file-l
        myPath = destPath + "/" + str(self.myYear)
        dirControl = os.path.isdir(myPath)
        if not dirControl:
            os.mkdir(myPath)

        self.saidiFile = myPath + "/" + str(self.myMonth) + " Deconectari.xlsx"
        fileControl = os.path.isfile(self.saidiFile)
        if not fileControl:
            myOriginal = os.path.abspath(".") + "/Bundle/Ungheni/Deconectari/Deconectari.xlsx"
            shutil.copyfile(myOriginal, self.saidiFile)

        self.wbDec = load_workbook(self.saidiFile)
        self.wsDecProg = self.wbDec["programate-MT+JT"]
        self.wsDecNeProg = self.wbDec["neprogramate-JT"]


    def dtRegAl(self):
        self.destLoad()
        destPath = self.wsDest.cell(row=1, column=2).value

        self.myMonth = datetime.datetime.now().strftime("%m")

        # Introduc datele in Registru cu controlul mapelor si file-l
        # myPath = "D:/Red-Nord/DISC T/Autorizatii/" + str(self.myYear)
        # dirControl = os.path.isdir(myPath)
        # if not dirControl:
        #     os.mkdir(myPath)
        if destPath != "":
            self.regFile = destPath + "/" + "Registru AUTORIZATIILOR UN " + str(self.myMonth) + ".xlsx"
            fielControl = os.path.isfile(self.regFile)
            if not fielControl:
                myOriginal = os.path.abspath(".") + "/Bundle/Ungheni/Registre/Registru AUTORIZATIILOR UN.xlsx"
                shutil.copyfile(myOriginal, self.regFile)

        self.wbRegAl = load_workbook(self.regFile)
        self.wsRegAl = self.wbRegAl["Registru"]
        self.wsRegAlLink = self.wbRegAl["alLink"]

    def dtAnAnual(self):
        self.myYear = datetime.datetime.now().strftime("%Y")
        # Introduc datele in Registru cu controlul mapelor si file-l
        myPath = os.path.abspath(".") + "/Bundle/Ungheni/Analiza/Anuala"
        dirControl = os.path.isdir(myPath)
        if not dirControl:
            os.makedirs(myPath)
        self.fileAnAnual = myPath + "/" + str(self.myYear) + " PT_Ungheni_analiza.xlsx"
        dirControl = os.path.isfile(self.fileAnAnual)
        if not dirControl:
            myOriginal = os.path.abspath(".") + "/Bundle/Ungheni/PT_Ungheni_analiza.xlsx"
            shutil.copyfile(myOriginal, self.fileAnAnual)
        self.wbAnAnual = load_workbook(self.fileAnAnual)
        self.wsAnAnualP = self.wbAnAnual["Programat"]
        self.wsAnAnualN = self.wbAnAnual["Neprogramat"]

    def ofChangeAn(self):
        # self.abrOficii()
        if not self.ofChangeContrSec:
            self.ofChangeContr = True
            self.analizaContr = True
            self.ofAfterCh = self.ofCombo.currentText()
            self.decAnaliza()
            self.ofChangeContrSec = False

    def decAnalizaLun(self):
        self.dtContrSaidi()
        #Controlez cind schimb oficiul
        self.ofChangeContr = False
        # Controlez cind schimb oficiul II
        self.ofChangeContrSec = False
        # Controlez analiza lunara
        self.analizaContr = True
        # Controlez analiza anuala
        self.analizaAnContr = 0
        self.decAnaliza()

    def decAnalizaAn(self):
        self.dtAnAnual()
        self.ofChangeContr = False
        self.ofChangeContrSec = False
        self.analizaContr = True
        self.analizaAnContr = 1
        self.decAnaliza()

    def decAnaliza(self):
        # self.abrOficii()
        self.ofCombo = QComboBox()
        self.ofCombo.addItems(self.ofListAbr)
        self.ofCombo.setStyleSheet("padding-left:10%; font-size:12px")
        self.ofCombo.setFixedHeight(25)
        self.ofCombo.setFixedWidth(100)
        # self.ofCombo.setEditable(True)
        self.ofCombo.currentTextChanged.connect(self.ofChangeAn)

        self.ptFidPl = []
        self.ptFidPlCons = []
        self.ptFidPlOre = []
        self.ptFidPlOf = []
        self.nrDec = []
        if self.analizaContr and self.analizaAnContr == 0:
            for i in range(19, self.wsDecProg.max_row + 1):
                self.ptFidPl.append(self.wsDecProg.cell(row=i, column=5).value + \
                                self.wsDecProg.cell(row=i, column=6).value)
                self.ptFidPlCons.append(self.wsDecProg.cell(row=i, column=15).value)
                self.ptFidPlOre.append(self.wsDecProg.cell(row=i, column=12).value)
                self.ptFidPlOf.append(self.wsDecProg.cell(row=i, column=2).value)
        elif not self.analizaContr and self.analizaAnContr == 0:
            for i in range(21, self.wsDecNeProg.max_row + 1):
                self.ptFidPl.append(self.wsDecNeProg.cell(row=i, column=4).value + \
                                 self.wsDecNeProg.cell(row=i, column=5).value)
                self.ptFidPlCons.append(self.wsDecNeProg.cell(row=i, column=27).value)
                self.ptFidPlOre.append(self.wsDecNeProg.cell(row=i, column=19).value)
                self.ptFidPlOf.append(self.wsDecNeProg.cell(row=i, column=2).value)
        elif self.analizaAnContr == 1:
            for i in range(2, self.wsAnAnualP.max_row + 1):
                self.ptFidPl.append(self.wsAnAnualP.cell(row=i, column=2).value)
                self.ptFidPlCons.append(self.wsAnAnualP.cell(row=i, column=3).value)
                self.ptFidPlOre.append(self.wsAnAnualP.cell(row=i, column=4).value)
                self.ptFidPlOf.append(self.wsAnAnualP.cell(row=i, column=1).value)
        elif self.analizaAnContr == 2:
            for i in range(2, self.wsAnAnualN.max_row + 1):
                self.ptFidPl.append(self.wsAnAnualN.cell(row=i, column=2).value)
                self.ptFidPlCons.append(self.wsAnAnualN.cell(row=i, column=3).value)
                self.ptFidPlOre.append(self.wsAnAnualN.cell(row=i, column=4).value)
                self.ptFidPlOf.append(self.wsAnAnualN.cell(row=i, column=1).value)

        for i in self.ptFidPl:
            self.nrDec.append(self.ptFidPl.count(i))

        self.anTabel = pd.DataFrame()
        for i in range(len(self.ptFidPl)):
            self.anTabel.at[i, 0] = self.ptFidPl[i]
            self.anTabel.at[i, 1] = self.ptFidPlCons[i]
            self.anTabel.at[i, 2] = self.nrDec[i]
            myList = self.ptFidPlOre[i].split(":")
            self.anTabel.at[i, 3] = int(myList[0]) + int(myList[1])/60
            self.anTabel.at[i, 4] = self.ptFidPlOf[i]

        try:
            self.anTabel.sort_values(by=0, ignore_index=True, inplace=True)
        except KeyError:
            if self.analizaContr and self.analizaAnContr == 0:
                self.msSecCall("Deconectari lunare programate nu exista!")
            elif not self.analizaContr and self.analizaAnContr == 0:
                self.msSecCall("Deconectari lunare neprogramate nu exista!")
            elif self.analizaAnContr == 1:
                self.msSecCall("Deconectari anuale programate nu exista!")
            elif self.analizaAnContr == 2:
                self.msSecCall("Deconectari anuale neprogramate nu exista!")

        iterIAn = []
        for i in range(1, len(self.ptFidPl)):
            if self.anTabel.at[i, 0] == self.anTabel.at[i - 1, 0]:
                self.anTabel.at[i, 3] = self.anTabel.at[i, 3] + \
                                            self.anTabel.at[i-1, 3]
                self.anTabel.drop([i-1], inplace=True)
        try:
            self.anTabel.sort_values(by=3, ignore_index=True, inplace=True, ascending=False)
        except KeyError:
            pass

        for i in range(len(self.anTabel)):
            if self.ofChangeContr:
                if self.anTabel.at[i, 4] != self.ofAfterCh:
                    iterIAn.append(i)
            else:
                if self.anTabel.at[i, 4] != self.ofCombo.currentText():
                    iterIAn.append(i)

        for i in range(len(iterIAn)):
            self.anTabel.drop([iterIAn[i]], inplace=True)

        try:
            self.anTabel.sort_values(by=3, ignore_index=True, inplace=True, ascending=False)
        except KeyError:
            pass

    #Pun graficile analizei

        categories = []
        set0 = QtCharts.QBarSet("Nr. de deconectari")
        set1 = QtCharts.QBarSet("Ore deconectate")
        set2 = QtCharts.QBarSet("Nr. de consumatori x10")

        for i in range(len(self.anTabel)):
            set0.append([float(self.anTabel.at[i, 2])])
            set1.append([float(self.anTabel.at[i, 3])])
            set2.append([float(self.anTabel.at[i, 1]) / 10])
            categories.append(self.anTabel.at[i, 0])

        series = QtCharts.QBarSeries()
        series.append(set0)
        series.append(set1)
        series.append(set2)

        chart = QtCharts.QChart()
        chart.addSeries(series)
        if self.analizaContr or self.analizaAnContr == 1:
            chart.setTitle("Deconectari programate")
        elif not self.analizaContr or self.analizaAnContr == 2:
            chart.setTitle("Deconectari neprogramate")
        chart.setTitleFont(QFont("Calibri", 14))
        # chart.setTitleBrush(QColor(191, 60, 60))
        chart.setAnimationOptions(QtCharts.QChart.SeriesAnimations)

        axis = QtCharts.QBarCategoryAxis()
        axis.append(categories)
        chart.createDefaultAxes()
        chart.setAxisX(axis, series)

        chart.legend().setVisible(True)
        chart.legend().setAlignment(Qt.AlignBottom)


        self.chartView = QtCharts.QChartView(chart)
        self.chartView.setRenderHint(QPainter.Antialiasing)
        self.anFuncWindow()

    def anFuncWindow(self):

        if self.analizaContr or self.analizaAnContr == 1:
            self.anFrameP = QFrame()
            vbox = QVBoxLayout()
            vbox.addWidget(self.chartView)
            self.anFrameP.setLayout(vbox)
            self.analizaContr = False
            if self.analizaAnContr == 1:
                self.analizaAnContr = 2
            self.decAnaliza()
        elif not self.analizaContr or self.analizaAnContr == 2:
            if self.analizaAnContr == 0:
                anLunTitle = QLabel("Analiza lunara de deconectari, oficiul:")
            elif self.analizaAnContr == 2:
                anLunTitle = QLabel("Analiza anuala de deconectari, oficiul:")
            anLunTitle.setStyleSheet("padding-left: 50%; font-size:24px; color:rgb(191, 60, 60)")
            self.anFrameN = QFrame()

            emptyLb = QLabel("")
            emptyLb2 = QLabel("")
            emptyLb3 = QLabel("")
            emptyLb4 = QLabel("")
            # emptyLb.setStyleSheet("margin-right:1300%")

            hbox = QHBoxLayout()
            hbox.addWidget(anLunTitle)
            hbox.addWidget(self.ofCombo)
            hbox.addWidget(emptyLb)
            hbox.addWidget(emptyLb2)
            hbox.addWidget(emptyLb3)
            hbox.addWidget(emptyLb4)
            hbox.setStretch(1, 1)

            vbox = QVBoxLayout()
            vbox.addLayout(hbox)
            vbox.addWidget(self.chartView)
            self.anFrameN.setLayout(vbox)

            myScrRes = app.primaryScreen()
            myScrAvailable = myScrRes.availableGeometry()
            myHeight = myScrAvailable.height()
            mySplitter = QSplitter(Qt.Vertical)
            mySplitter.addWidget(self.anFrameN)
            mySplitter.addWidget(self.anFrameP)
            mySplitter.handle(1).setStyleSheet("Background-color: rgb(191, 60, 60)")
            mySplitter.setSizes([int(myHeight/5), int(myHeight/5)])

            if self.anControl:
                self.anWindow.close()

            self.anWindow = QMdiSubWindow()
            self.myMidi.addSubWindow(self.anWindow)
            self.anWindow.setWindowIcon(QIcon(QPixmap(1, 1)))
            self.anWindow.setWidget(mySplitter)
            self.anWindow.setGeometry(100, 100, 1000, 600)
            self.anControl = True
            if self.ofChangeContr:
                self.ofChangeContrSec = True
                self.ofCombo.setCurrentText(self.ofAfterCh)
            self.anWindow.showMaximized()
            # self.anWindow.show()

    def deranjFunc(self):
        if self.ptLine.text() == "PT":
            self.ptLine.setText("")
        if self.ptFidLine.text() == "Fider nr.":
            self.ptFidLine.setText("")
        self.abrOficii()
        pyDateTime = datetime.datetime.now()
        self.db.deranjamente.insert_one({
            "oficiul": self.ofVar,
            "nr_ordine": str(self.db.deranjamente.estimated_document_count() + 1),
            "transmis": self.sfLine.text(),
            "sector": self.sectCombo.currentText(),
            "instalatia": self.instalatiaCombo.currentText(),
            "fid_10kv": self.f10kvLine.text(),
            "pt": self.ptLine.text(),
            "fid_04kv": self.ptFidLine.text(),
            "continutul": self.continText.toPlainText() + "!",
            "data": pyDateTime.strftime("%d.%m.%Y %H:%M"),
            "responsabil": "-//-",
            "starea": "Neexecutat"
        })
        self.dialBox.close()
        self.deranjPop()

    def deranjPop(self):
        if self.deranjControlPop == True:
            self.tabWindowDeranj.close()

        self.data = pd.DataFrame(self.db.deranjamente.find({}, {"_id": 0}))
        self.data.columns = range(12)
        # print(self.data)
        self.data.sort_index(ascending=False, inplace=True, ignore_index=True)
        header = ["Oficiul", "Nr.", "Transmis", "Sectorul", "Instalatia", "Fider 10kV", "PT", "Fider 0,4kV",
                  "Continutul", "Data, ora", "Semnatura, Responsabil", "Starea"]

        self.table = QTableView()
        self.model = TableModel(self.data, header)
        self.table.setModel(self.model)
        self.table.setWordWrap(True)
        self.table.setTextElideMode(Qt.ElideMiddle)
        self.table.resizeColumnsToContents()
        self.table.setStyleSheet("Background-color: rgb(200, 200, 200)")
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.verticalHeader().hide()
        self.table.setColumnWidth(0, 100)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(2, 200)
        self.table.setColumnWidth(3, 100)
        self.table.setColumnWidth(4, 150)
        self.table.setColumnWidth(5, 150)
        self.table.setColumnWidth(6, 150)
        self.table.setColumnWidth(7, 100)
        self.table.setColumnWidth(8, 250)
        self.table.setColumnWidth(11, 250)
        self.table.resizeRowsToContents()

        regFrame = QFrame()
        regAlTitle = QLabel()

        regAlTitle.setText("Registru de deranjamente Of.:")
        regAlTitle.setStyleSheet("padding-left: 50%; font-size:24px; color:rgb(191, 60, 60)")
        # regAlTitle.move(0, 100)
        # self.loadOficii()
        myList = ['Toate']
        myList = myList + self.ofList
        self.ofCombo = QComboBox()
        self.ofCombo.addItems(myList)
        # self.ofComboReg.setStyleSheet("margin-right:1400%")
        self.ofCombo.setStyleSheet("padding-left:10%; font-size:12px")
        self.ofCombo.setFixedHeight(25)
        self.ofCombo.setFixedWidth(100)
        self.ofCombo.currentTextChanged.connect(self.ofChangeReg)


        #Combo pentru sectoare
        self.sectCombo = QComboBox()
        self.sectorList = []
        self.sectCombo.addItems(self.sectorList)
        self.sectCombo.setStyleSheet("padding-left:10%; font-size:12px")
        self.sectCombo.setFixedHeight(25)
        self.sectCombo.setFixedWidth(100)
        self.sectCombo.textActivated.connect(self.changeSect)

        self.execCheck = QCheckBox("Neexecutat")
        self.execCheck.setStyleSheet("padding-left:10%; font-size:12px")
        self.execCheck.stateChanged.connect(self.changeExec)

        # emptyLb = QLabel("")
        # emptyLb2 = QLabel("")
        emptyLb3 = QLabel("")
        emptyLb4 = QLabel("")
        # emptyLb.setStyleSheet("margin-right:1400%")

        hbox = QHBoxLayout()
        hbox.addWidget(regAlTitle)
        hbox.addWidget(self.ofCombo)
        hbox.addWidget(self.sectCombo)
        hbox.addWidget(self.execCheck)
        hbox.addWidget(emptyLb3)
        hbox.addWidget(emptyLb4)
        hbox.setStretch(1, 1)

        vbox = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addWidget(self.table)
        regFrame.setLayout(vbox)

        self.tabWindowDeranj = QMdiSubWindow()
        self.myMidi.addSubWindow(self.tabWindowDeranj)
        self.tabWindowDeranj.setWindowIcon(QIcon(QPixmap(1, 1)))
        self.tabWindowDeranj.setWidget(regFrame)
        self.tabWindowDeranj.setGeometry(100, 100, 1000, 600)
        self.tabWindowDeranj.showMaximized()
        self.deranjControlPop = True
        self.tabWindowDeranj.show()


    def decFunc(self):

        self.dtContrDecZl()
        self.dtAnAnual()
        self.abrOficii()

        myMaxRow = self.wsDecPT.max_row + 1

        if self.wsDecPT.cell(row=2, column=8).value == None:
            self.wsDecPT.cell(row=2, column=8).value = datetime.datetime.now().strftime("%d.%m.%Y")
        self.wsDecPT.cell(row=myMaxRow, column=1).value = \
            str(int(self.wsDecPT.cell(row=myMaxRow - 1, column=1).value) + 1)
        self.wsDecPT.cell(row=myMaxRow, column=1).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecPT.cell(row=myMaxRow, column=2).value = self.ofVar
        self.wsDecPT.cell(row=myMaxRow, column=2).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecPT.cell(row=myMaxRow, column=3).value = self.ptLine.text()
        self.wsDecPT.cell(row=myMaxRow, column=4).value = self.ptFidLine.text()
        self.wsDecPT.cell(row=myMaxRow, column=4).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecPT.cell(row=myMaxRow, column=5).value = self.dtLine.text()
        self.wsDecPT.cell(row=myMaxRow, column=5).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecPT.cell(row=myMaxRow, column=6).value = datetime.datetime.now().strftime("%d.%m.%y %H:%M")
        self.wsDecPT.cell(row=myMaxRow, column=6).alignment = \
            Alignment(horizontal="center", vertical="center")
        # Calculez si pun orele diferenta lor
        valueDate_6 = self.wsDecPT.cell(row=myMaxRow, column=6).value
        valueDate_5 = self.wsDecPT.cell(row=myMaxRow, column=5).value
        strToDate_6 = datetime.datetime.strptime(valueDate_6, "%d.%m.%y %H:%M")
        strToDate_5 = datetime.datetime.strptime(valueDate_5, "%d.%m.%y %H:%M")
        delta_6_5 = strToDate_6 - strToDate_5
        myList = str(delta_6_5).split()
        try:
            if int(myList[0]):
                myDateSplit = myList[2].split(":")
                myDayPlus = 24 * int(myList[0]) + int(myDateSplit[0])
                myDateStr = str(myDayPlus) + ":" + myDateSplit[1] + ":" + myDateSplit[2]
                myDeltaHour = myDateStr
        except ValueError:
            myDeltaHour = str(delta_6_5)
        self.wsDecPT.cell(row=myMaxRow, column=7).value = myDeltaHour
        self.wsDecPT.cell(row=myMaxRow, column=7).alignment = \
            Alignment(horizontal="center", vertical="center")

        # Calculez si pun numarul de consumatori si localitatea
        for i in range(2, self.wsPt.max_row + 1):
            if self.wsPt.cell(row=i, column=1).value == self.wsDecPT.cell(row=myMaxRow, column=3).value:
                self.totNrCas = int(self.wsPt.cell(row=i, column=4).value)
                self.fazaNrCas = round(self.totNrCas / 9)
                if self.fazaNrCas > 25:
                    self.fazaNrCas = random.randrange(20, 30)
                self.wsDecPT.cell(row=myMaxRow, column=8).value = str(self.fazaNrCas)
                self.wsDecPT.cell(row=myMaxRow, column=8).alignment = \
                    Alignment(horizontal="center", vertical="center")

                self.totNrEc = self.wsPt.cell(row=i, column=5).value
                if self.totNrEc <= 2 and self.totNrCas == 0:
                    self.fazaNrEc = 1
                elif self.totNrEc > 12:
                    self.fazaNrEc = random.randrange(1, 3)
                else:
                    self.fazaNrEc = round(self.totNrEc / 9)
                self.wsDecPT.cell(row=myMaxRow, column=9).value = str(self.fazaNrEc)
                self.wsDecPT.cell(row=myMaxRow, column=9).alignment = \
                    Alignment(horizontal="center", vertical="center")
                self.wsDecPT.cell(row=myMaxRow, column=11).value = \
                    self.wsPt.cell(row=i, column=2).value
                myLocalitate = self.wsPt.cell(row=i, column=2).value

        self.wsDecPT.cell(row=myMaxRow, column=10).value = \
            str(int(self.wsDecPT.cell(row=myMaxRow, column=8).value) + \
            int(self.wsDecPT.cell(row=myMaxRow, column=9).value))
        self.wsDecPT.cell(row=myMaxRow, column=10).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecPT.cell(row=myMaxRow, column=12).value = self.cauzaCombo.currentText()
        self.wsDecPT.cell(row=myMaxRow, column=13).value = self.termenCombo.currentText()
        self.wsDecPT.cell(row=myMaxRow, column=13).alignment = \
            Alignment(horizontal="center", vertical="center")
        try:
            self.wbDecZil.save(self.rapFile)
        except PermissionError:
            self.msSecCall("Datele nu s-au introdus in fisierul RAPORT PDJT excel \n"
                           "(cineva foloseste aplicatia)!")

        self.dtContrSaidi()
        myMaxRow = self.wsDecNeProg.max_row + 1
        self.wsDecNeProg.cell(row=myMaxRow, column=1).value = \
            self.wsDecNeProg.cell(row=myMaxRow - 1, column=1).value + 1
        self.wsDecNeProg.cell(row=myMaxRow, column=1).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=2).value = \
            self.ofVar
        self.wsDecNeProg.cell(row=myMaxRow, column=2).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=3).value = \
            datetime.date.today().strftime("%d.%m.%y")
        self.wsDecNeProg.cell(row=myMaxRow, column=3).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=4).value = \
            self.ptLine.text()
        self.wsDecNeProg.cell(row=myMaxRow, column=4).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=5).value = \
            self.ptFidLine.text()
        self.wsDecNeProg.cell(row=myMaxRow, column=5).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=6).value = \
            0
        self.wsDecNeProg.cell(row=myMaxRow, column=6).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=7).value = \
            0
        self.wsDecNeProg.cell(row=myMaxRow, column=7).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=8).value = \
            1
        self.wsDecNeProg.cell(row=myMaxRow, column=8).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=9).value = \
            1
        self.wsDecNeProg.cell(row=myMaxRow, column=9).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=10).value = \
            self.dtLine.text()
        self.wsDecNeProg.cell(row=myMaxRow, column=10).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=11).value = \
            datetime.datetime.now().strftime("%d.%m.%y %H:%M")
        self.wsDecNeProg.cell(row=myMaxRow, column=11).alignment = \
            Alignment(horizontal="center", vertical="center")

        self.wsDecNeProg.cell(row=myMaxRow, column=19).value = \
            myDeltaHour
        self.wsDecNeProg.cell(row=myMaxRow, column=19).alignment = \
            Alignment(horizontal="center", vertical="center")

        self.wsDecNeProg.cell(row=myMaxRow, column=25).value = self.fazaNrCas
        self.wsDecNeProg.cell(row=myMaxRow, column=25).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=26).value = self.fazaNrEc
        self.wsDecNeProg.cell(row=myMaxRow, column=26).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=27).value = \
            self.wsDecNeProg.cell(row=myMaxRow, column=25).value + \
            self.wsDecNeProg.cell(row=myMaxRow, column=26).value
        self.wsDecNeProg.cell(row=myMaxRow, column=27).alignment = \
            Alignment(horizontal="center", vertical="center")
        self.wsDecNeProg.cell(row=myMaxRow, column=28).value = \
            myLocalitate

        self.wsDecNeProg.cell(row=myMaxRow, column=32).value = \
            self.cauzaCombo.currentText()

        try:
            self.wbDec.save(self.saidiFile)
        except PermissionError:
            self.msSecCall("Datele nu s-au introdus in fisierul SAIDI excel (cineva foloseste aplicatia)!")

        # Introduc datele in Excel analiza anuala
        anAnualMaxRow = self.wsAnAnualN.max_row + 1
        self.wsAnAnualN.cell(row=anAnualMaxRow, column=1).value = \
            self.ofVar
        self.wsAnAnualN.cell(row=anAnualMaxRow, column=2).value = \
            self.ptLine.text() + " F" + self.ptFidLine.text()
        self.wsAnAnualN.cell(row=anAnualMaxRow, column=3).value = \
            self.wsDecNeProg.cell(row=self.wsDecNeProg.max_row, column=25).value + \
            self.wsDecNeProg.cell(row=self.wsDecNeProg.max_row, column=26).value
        self.wsAnAnualN.cell(row=anAnualMaxRow, column=4).value = \
            myDeltaHour
        try:
            self.wbAnAnual.save(self.fileAnAnual)
        except PermissionError:
            self.msSecCall("Datele nu vor participa la analiza anuala,\n"
                           "(undeva este deschisa analiza anuala excel, nu se permite introducerea datelor)!")

        self.dialBox.close()
        self.decTabWindow()

    def okPass(self):
        try:
            i = self.angajati.find_one({"name": self.uCombo.currentText()})
            print(self.uCombo.currentText())
            if i["nr_tabel"] == self.psText.text():
                self.dialInt.close()
                global closeApp
                closeApp = False
            else:
                self.msSecCall("Nu ati introdus corect parola sau userul!")
            self.passControl = True
        except TypeError:
            self.msSecCall("Nu ati introdus corect parola sau userul!")

    def cancelPass(self):
        self.dialInt.close()
        self.close()
        global closeApp
        closeApp = True

    def logOut(self):
        self.passControl = False
        self.intTrig()


    def okTrigger(self):
        self.okControl = False
        if self.sfLine.text() == "":
            self.msCall('"Sef de lucrari:"')
            self.sfLine.setFocus()
        elif self.emLine.text() == "":
            self.msCall('"Emitent:"')
            self.emLine.setFocus()
        elif self.instLine.currentText() == "":
            self.msCall('"Se executa (tipul instalatiei (LEA, ID etc.)):"')
            self.instLine.setFocus()
        elif self.lucrLine.currentText() == "Lucrarile efectuate:"\
                or self.lucrLine.currentText() == "":
            self.msCall('"Lucrarile efectuate:"')
            self.lucrLine.setFocus()
        elif self.decCombo.currentText() == "Programat" or self.decCombo.currentText() == "Neprogramat":
            if self.ptLine.text() == "PT" or self.ptLine.text() == "":
                self.msCall('"PT"')
                self.ptLine.setFocus()
            elif self.ptFidLine.text() == "Fider nr.":
                self.msCall('"Fider nr."')
                self.ptFidLine.setFocus()
            elif self.smNr.text() == "SM nr.":
                self.msCall('"SM nr."')
                self.smNr.setFocus()
            else:
                self.regPop()
        else:
            self.regPop()

    def okDsTrigger(self):
        self.okControl = True
        if self.sfLine.text() == "":
            self.msCall('"Sef de lucrari:"')
            self.sfLine.setFocus()
        # elif self.memEchLine.text() == "":
        #     self.msCall('"Membrii echipei:"')
        #     self.memEchLine.setFocus()
        elif self.emLine.text() == "":
            self.msCall('"Emitent:"')
            self.emLine.setFocus()
        # elif self.instLine.text() == "Instalatia:":
        #     self.msCall('"Instalatia:"')
        #     self.instLine.setFocus()
        # elif self.ptLine.text() == "PT" or self.ptLine.text() == "":
        #     self.msCall('"PT:"')
        #     self.ptLine.setFocus()
        elif self.lucrLine.currentText() == "Lucrarile efectuate:" \
                or self.lucrLine.currentText() == "":
            self.msCall('"Lucrarile efectuate:"')
            self.lucrLine.setFocus()
        else:
            self.regPop()

    def okDecTrigger(self):
        if self.ptLine.text() == "PT" or self.ptLine.text() == "":
            self.msCall('"PT:"')
            self.ptLine.setFocus()
        elif self.ptFidLine.text() == "Fider nr.":
            self.msCall('"Fider nr.')
            self.ptFidLine.setFocus()
        else:
            self.decFunc()

    def okDeranjTrigger(self):
        if self.sectCombo.currentText() == "Alege:" or self.sectCombo.currentText() == "":
            self.msCall('"Sectorul:"')
            self.sectCombo.setFocus()
        elif self.instalatiaCombo.currentText() == "Alege:" or self.instalatiaCombo.currentText() == "":
            self.msCall('"Instalatia:"')
            self.instalatiaCombo.setFocus()
        elif (self.instalatiaCombo.currentText() == "LEA-0.4kV" or \
                self.instalatiaCombo.currentText() == "ID-10kV" or \
                self.instalatiaCombo.currentText() == "ID-0.4kV" or \
                self.instalatiaCombo.currentText() == "Consumator casnic") and \
                (self.ptLine.text() == "PT" or self.ptLine.text() == ""):
            self.msCall('"PT"')
            self.ptLine.setFocus()
        elif (self.instalatiaCombo.currentText() == "LEA-10kV" or \
                self.instalatiaCombo.currentText() == "LC-10kV") and \
                self.f10kvLine.text() == "":
            self.msCall('"Fid.10kV nr:"')
            self.f10kvLine.setFocus()
        elif self.continText.toPlainText() == "":
            self.msCall('"Continutul:"')
            self.continText.setFocus()
        else:
            self.deranjFunc()


app = QApplication(sys.argv)

#Making a splash screen
pixmap = QPixmap("ataman_splash.png")
splash = QSplashScreen(pixmap)
splash.show()
splash.showMessage("Loading...")
app.processEvents()

mw = mainWindow()
try:
    if not closeApp:
        mw.show()
        splash.finish(mw)
        app.exec_()
except NameError:
    pass